#!/usr/bin/env python3
"""
╔═══════════════════════════════════════════════════════════════════════╗
║  ▓▓▓ SYNTHGRID ▓▓▓  AI-POWERED EXCEL AUTOMATION NODE                ║
║  Target: Linux/Fedora/KDE | Engine: PyQt6 + Ollama/Groq              ║
╚═══════════════════════════════════════════════════════════════════════╝
"""

import sys
import os
import json
import time
import threading
import io
import traceback
from datetime import datetime
from typing import Optional, Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, 
    GradientFill, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.datavalidation import DataValidation

import pandas as pd

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTextEdit, QLineEdit, QLabel, QPushButton, QSplitter, QFrame,
    QTableWidget, QTableWidgetItem, QScrollArea, QComboBox,
    QStatusBar, QHeaderView, QTabWidget, QFileDialog, QProgressBar,
    QCheckBox, QGroupBox, QGridLayout, QSpinBox, QSlider
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QTimer, QPropertyAnimation, 
    QEasingCurve, pyqtProperty, QRect
)
from PyQt6.QtGui import (
    QFont, QColor, QPalette, QTextCharFormat, QSyntaxHighlighter,
    QBrush, QPainter, QPen, QLinearGradient, QFontMetrics,
    QKeySequence, QShortcut, QIcon, QPixmap
)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL ENGINE
# ─────────────────────────────────────────────────────────────────────────────

class ExcelEngine:
    """Core Excel manipulation engine — MCP parity tool set"""
    
    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.filepath: Optional[str] = None
        self.active_sheet_name: Optional[str] = None
        self._new_workbook()

    def _new_workbook(self):
        self.workbook = Workbook()
        ws = self.workbook.active
        ws.title = "Sheet1"
        self.active_sheet_name = "Sheet1"
        self.filepath = None

    def _get_sheet(self, sheet_name: Optional[str] = None):
        name = sheet_name or self.active_sheet_name
        if name and name in self.workbook.sheetnames:
            return self.workbook[name]
        return self.workbook.active

    # ── Workbook CRUD ──────────────────────────────────────────────────────

    def create_workbook(self, filepath: str) -> dict:
        self._new_workbook()
        self.filepath = filepath
        self.workbook.save(filepath)
        return {"status": "ok", "message": f"Workbook created: {filepath}"}

    def open_workbook(self, filepath: str) -> dict:
        if not os.path.exists(filepath):
            return {"status": "error", "message": f"File not found: {filepath}"}
        self.workbook = openpyxl.load_workbook(filepath)
        self.filepath = filepath
        self.active_sheet_name = self.workbook.sheetnames[0]
        return {"status": "ok", "message": f"Opened: {filepath}", "sheets": self.workbook.sheetnames}

    def save_workbook(self, filepath: Optional[str] = None) -> dict:
        path = filepath or self.filepath
        if not path:
            return {"status": "error", "message": "No filepath specified"}
        self.workbook.save(path)
        self.filepath = path
        return {"status": "ok", "message": f"Saved: {path}"}

    def get_workbook_info(self) -> dict:
        if not self.workbook:
            return {"status": "error", "message": "No workbook open"}
        info = {"sheets": [], "filepath": self.filepath}
        for name in self.workbook.sheetnames:
            ws = self.workbook[name]
            info["sheets"].append({
                "name": name,
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_col": ws.max_column
            })
        return {"status": "ok", "info": info}

    # ── Sheet Management ───────────────────────────────────────────────────

    def create_sheet(self, name: str, position: Optional[int] = None) -> dict:
        if name in self.workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{name}' already exists"}
        ws = self.workbook.create_sheet(title=name, index=position)
        self.active_sheet_name = ws.title
        return {"status": "ok", "message": f"Sheet '{name}' created"}

    def delete_sheet(self, name: str) -> dict:
        if name not in self.workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{name}' not found"}
        del self.workbook[name]
        if self.workbook.sheetnames:
            self.active_sheet_name = self.workbook.sheetnames[0]
        return {"status": "ok", "message": f"Sheet '{name}' deleted"}

    def rename_sheet(self, old_name: str, new_name: str) -> dict:
        if old_name not in self.workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{old_name}' not found"}
        self.workbook[old_name].title = new_name
        if self.active_sheet_name == old_name:
            self.active_sheet_name = new_name
        return {"status": "ok", "message": f"Renamed '{old_name}' → '{new_name}'"}

    def copy_sheet(self, source: str, target: str) -> dict:
        if source not in self.workbook.sheetnames:
            return {"status": "error", "message": f"Source sheet '{source}' not found"}
        ws = self.workbook.copy_worksheet(self.workbook[source])
        ws.title = target
        return {"status": "ok", "message": f"Copied '{source}' → '{target}'"}

    def set_active_sheet(self, name: str) -> dict:
        if name not in self.workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{name}' not found"}
        self.active_sheet_name = name
        return {"status": "ok", "message": f"Active sheet: '{name}'"}

    # ── Cell Operations ────────────────────────────────────────────────────

    def read_cell(self, cell: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        val = ws[cell].value
        return {"status": "ok", "cell": cell, "value": val, "sheet": ws.title}

    def write_cell(self, cell: str, value: Any, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws[cell] = value
        return {"status": "ok", "message": f"Written {value!r} → {cell}"}

    def write_range(self, start_cell: str, data: list, sheet: Optional[str] = None) -> dict:
        """Write 2D list starting from start_cell"""
        ws = self._get_sheet(sheet)
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_str, row = coordinate_from_string(start_cell)
        col = column_index_from_string(col_str)
        for r_idx, row_data in enumerate(data):
            for c_idx, val in enumerate(row_data):
                ws.cell(row=row + r_idx, column=col + c_idx, value=val)
        return {"status": "ok", "message": f"Written {len(data)}×{len(data[0]) if data else 0} range from {start_cell}"}

    def read_range(self, cell_range: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        data = []
        for row in ws[cell_range]:
            data.append([cell.value for cell in row])
        return {"status": "ok", "range": cell_range, "data": data}

    def set_formula(self, cell: str, formula: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws[cell] = formula if formula.startswith("=") else f"={formula}"
        return {"status": "ok", "message": f"Formula set at {cell}: {ws[cell].value}"}

    def delete_row(self, row: int, count: int = 1, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.delete_rows(row, count)
        return {"status": "ok", "message": f"Deleted {count} row(s) from row {row}"}

    def insert_row(self, row: int, count: int = 1, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.insert_rows(row, count)
        return {"status": "ok", "message": f"Inserted {count} row(s) at row {row}"}

    def delete_column(self, col: str, count: int = 1, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        col_idx = column_index_from_string(col)
        ws.delete_cols(col_idx, count)
        return {"status": "ok", "message": f"Deleted {count} column(s) from {col}"}

    def insert_column(self, col: str, count: int = 1, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        col_idx = column_index_from_string(col)
        ws.insert_cols(col_idx, count)
        return {"status": "ok", "message": f"Inserted {count} column(s) at {col}"}

    # ── Formatting ────────────────────────────────────────────────────────

    def format_cells(self, cell_range: str, options: dict, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        for row in ws[cell_range]:
            for cell in row:
                if "font_bold" in options or "font_size" in options or "font_color" in options or "font_name" in options:
                    existing = cell.font or Font()
                    cell.font = Font(
                        bold=options.get("font_bold", existing.bold),
                        size=options.get("font_size", existing.size),
                        color=options.get("font_color", "000000"),
                        name=options.get("font_name", existing.name or "Calibri"),
                        italic=options.get("font_italic", existing.italic),
                        underline=options.get("font_underline", existing.underline),
                    )
                if "bg_color" in options:
                    cell.fill = PatternFill(
                        start_color=options["bg_color"].lstrip("#"),
                        end_color=options["bg_color"].lstrip("#"),
                        fill_type="solid"
                    )
                if "alignment" in options:
                    al = options["alignment"]
                    cell.alignment = Alignment(
                        horizontal=al.get("horizontal", "general"),
                        vertical=al.get("vertical", "center"),
                        wrap_text=al.get("wrap_text", False)
                    )
                if "border" in options:
                    b = options["border"]
                    style = b.get("style", "thin")
                    color = b.get("color", "000000")
                    side = Side(style=style, color=color)
                    cell.border = Border(left=side, right=side, top=side, bottom=side)
                if "number_format" in options:
                    cell.number_format = options["number_format"]
        return {"status": "ok", "message": f"Formatted range {cell_range}"}

    def set_column_width(self, col: str, width: float, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.column_dimensions[col].width = width
        return {"status": "ok", "message": f"Column {col} width = {width}"}

    def set_row_height(self, row: int, height: float, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.row_dimensions[row].height = height
        return {"status": "ok", "message": f"Row {row} height = {height}"}

    def merge_cells(self, cell_range: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.merge_cells(cell_range)
        return {"status": "ok", "message": f"Merged {cell_range}"}

    def unmerge_cells(self, cell_range: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.unmerge_cells(cell_range)
        return {"status": "ok", "message": f"Unmerged {cell_range}"}

    def add_conditional_formatting(self, cell_range: str, rule_type: str, options: dict, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        if rule_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min", start_color=options.get("start_color", "FF0000"),
                end_type="max",   end_color=options.get("end_color",   "00FF00")
            )
        elif rule_type == "cell_is":
            fill = PatternFill(start_color=options.get("color", "FFFF00"), fill_type="solid")
            rule = CellIsRule(
                operator=options.get("operator", "greaterThan"),
                formula=[str(options.get("value", 0))],
                fill=fill
            )
        else:
            return {"status": "error", "message": f"Unknown rule type: {rule_type}"}
        ws.conditional_formatting.add(cell_range, rule)
        return {"status": "ok", "message": f"Conditional format added to {cell_range}"}

    def add_data_validation(self, cell_range: str, validation_type: str, options: dict, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        dv = DataValidation(
            type=validation_type,
            formula1=str(options.get("formula1", "")),
            formula2=str(options.get("formula2", "")) if options.get("formula2") else None,
            allow_blank=options.get("allow_blank", True),
            showErrorMessage=options.get("show_error", True),
            errorTitle=options.get("error_title", "Invalid"),
            error=options.get("error_msg", "Value is not allowed"),
        )
        dv.sqref = cell_range
        ws.add_data_validation(dv)
        return {"status": "ok", "message": f"Data validation added to {cell_range}"}

    def add_autofilter(self, cell_range: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.auto_filter.ref = cell_range
        return {"status": "ok", "message": f"AutoFilter applied to {cell_range}"}

    def freeze_panes(self, cell: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.freeze_panes = cell
        return {"status": "ok", "message": f"Panes frozen at {cell}"}

    # ── Charts ─────────────────────────────────────────────────────────────

    def add_chart(self, chart_type: str = "bar", data_range: str = None, title: str = "Chart",
                  position: str = "E1", sheet: Optional[str] = None, chart_data: dict = None) -> dict:
        ws = self._get_sheet(sheet)
        
        # Handle AI's incorrect parameter format (chart_data with headers/rows)
        if chart_data is not None and data_range is None:
            headers = chart_data.get("headers", [])
            rows = chart_data.get("rows", [])
            if headers and rows:
                # Write chart data to sheet starting at A20 (leave space for original data)
                start_row = 20
                for c_idx, header in enumerate(headers, 1):
                    ws.cell(row=start_row, column=c_idx, value=header)
                for r_idx, row in enumerate(rows, start_row + 1):
                    for c_idx, val in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=val)
                # Determine data range - build full reference
                end_row = start_row + len(rows)
                end_col = len(headers)
                from openpyxl.utils import get_column_letter
                data_range = f"{get_column_letter(1)}{start_row}:{get_column_letter(end_col)}{end_row}"
                title = title or "Chart"
        
        # Validate required parameters
        if not data_range:
            return {"status": "error", "message": "data_range is required (e.g., 'A1:D10')"}
        
        chart_map = {"bar": BarChart, "line": LineChart, "scatter": ScatterChart, "Line": LineChart}
        ChartClass = chart_map.get(chart_type.lower(), BarChart)
        chart = ChartClass()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Category"

        parts = data_range.split(":")
        if len(parts) == 2:
            from openpyxl.utils.cell import column_index_from_string
            start_cell, end_cell = parts[0], parts[1]
            start_col, start_row = column_index_from_string(''.join(c for c in start_cell if c.isalpha())), int(''.join(c for c in start_cell if c.isdigit()))
            end_col, end_row = column_index_from_string(''.join(c for c in end_cell if c.isalpha())), int(''.join(c for c in end_cell if c.isdigit()))
            ref = Reference(ws, min_col=start_col, min_row=start_row, max_col=end_col, max_row=end_row)
            chart.add_data(ref, titles_from_data=True)
        chart.shape = 4
        ws.add_chart(chart, position)
        return {"status": "ok", "message": f"{chart_type.capitalize()} chart '{title}' added at {position}"}

    # ── Pivot / Summary ───────────────────────────────────────────────────

    def create_pivot_summary(self, source_range: str, target_cell: str,
                              row_field: int = 0, value_field: int = 1,
                              sheet: Optional[str] = None, target_sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        data = []
        for row in ws[source_range]:
            data.append([cell.value for cell in row])
        if not data or len(data) < 2:
            return {"status": "error", "message": "Insufficient data for pivot"}
        headers = data[0]
        rows_data = data[1:]
        pivot = {}
        for row in rows_data:
            if len(row) > max(row_field, value_field):
                key = row[row_field]
                val = row[value_field]
                try:
                    val = float(val) if val is not None else 0
                except (TypeError, ValueError):
                    val = 0
                pivot[key] = pivot.get(key, 0) + val
        target_ws = self._get_sheet(target_sheet)
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        col_str, start_row = coordinate_from_string(target_cell)
        start_col = column_index_from_string(col_str)
        target_ws.cell(row=start_row, column=start_col, value=headers[row_field] if row_field < len(headers) else "Key")
        target_ws.cell(row=start_row, column=start_col + 1, value=headers[value_field] if value_field < len(headers) else "Sum")
        for i, (key, val) in enumerate(pivot.items(), 1):
            target_ws.cell(row=start_row + i, column=start_col, value=key)
            target_ws.cell(row=start_row + i, column=start_col + 1, value=val)
        return {"status": "ok", "message": f"Pivot summary created at {target_cell} with {len(pivot)} groups"}

    # ── Data Read ─────────────────────────────────────────────────────────

    def get_sheet_data(self, sheet: Optional[str] = None, max_rows: int = 100) -> list:
        ws = self._get_sheet(sheet)
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))
        return rows

    def get_sheet_names(self) -> list:
        return self.workbook.sheetnames if self.workbook else []

    # ── Smart Operations ──────────────────────────────────────────────────────

    def sort_range(self, cell_range: str, key_column: int = 0, ascending: bool = True, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        data = []
        for row in ws[cell_range]:
            data.append([cell.value for cell in row])
        if not data or len(data) < 2:
            return {"status": "error", "message": "Insufficient data to sort"}
        header = data[0]
        rows = data[1:]
        rows.sort(key=lambda x: x[key_column] if key_column < len(x) else "", reverse=not ascending)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row):
                ws.cell(row=r_idx+1, column=c_idx+1, value=val)
        return {"status": "ok", "message": f"Sorted {len(rows)} rows by column {key_column+1}"}

    def add_totals_row(self, cell_range: str, label: str = "Total", sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        data = []
        for row in ws[cell_range]:
            data.append([cell.value for cell in row])
        if not data or len(data) < 2:
            return {"status": "error", "message": "Insufficient data"}
        num_cols = len(data[0])
        start_row = ws[cell_range].row
        start_col = ws[cell_range].column
        total_row = start_row + len(data)
        ws.cell(row=total_row, column=1, value=label)
        for col in range(2, num_cols + 1):
            col_letter = get_column_letter(col)
            ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}{start_row+1}:{col_letter}{total_row-1})")
        return {"status": "ok", "message": f"Added totals row at row {total_row}"}

    def apply_header_style(self, row: int = 1, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        return {"status": "ok", "message": f"Applied header style to row {row}"}

    def apply_alternating_rows(self, start_row: int = 2, sheet: Optional[str] = None) -> dict:
        ws = self.get_sheet(sheet)
        for row in range(start_row, ws.max_row + 1):
            color = "F2F2F2" if (row - start_row) % 2 == 0 else "FFFFFF"
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        return {"status": "ok", "message": "Applied alternating row colors"}

    def auto_size_columns(self, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        return {"status": "ok", "message": "Auto-sized columns"}

    def add_data_bars(self, cell_range: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        from openpyxl.formatting.rule import DataBarRule
        rule = DataBarRule(startType="min", endType="max", color="638EC6", showValue=True)
        ws.conditional_formatting.add(cell_range, rule)
        return {"status": "ok", "message": f"Added data bars to {cell_range}"}

    def highlight_top_values(self, cell_range: str, n: int = 10, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        rule = Rule(type="top10", rank=n, percent=False, fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        ws.conditional_formatting.add(cell_range, rule)
        return {"status": "ok", "message": f"Highlighted top {n} values in {cell_range}"}

    def add_status_column(self, threshold_col: int, threshold_value: float, new_col_name: str = "Status", sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        max_row = ws.max_row
        max_col = ws.max_column + 1
        new_col_letter = get_column_letter(max_col)
        ws.cell(row=1, column=max_col, value=new_col_name)
        for row in range(2, max_row + 1):
            cell_ref = f"${{{get_column_letter(threshold_col)}}}{row}"
            formula = f'=IF({get_column_letter(threshold_col)}{row}>={threshold_value},"Good","Review")'
            ws.cell(row=row, column=max_col, value=formula)
        return {"status": "ok", "message": f"Added status column '{new_col_name}'"}

    def create_summary_sheet(self, source_sheet: str, summary_name: str = "Summary", sheet: Optional[str] = None) -> dict:
        src_ws = self.workbook[source_sheet]
        if source_sheet in self.workbook.sheetnames:
            self.workbook.create_sheet(title=summary_name)
        else:
            return {"status": "error", "message": f"Source sheet '{source_sheet}' not found"}
        sum_ws = self.workbook[summary_name]
        sum_ws["A1"] = "Column"
        sum_ws["B1"] = "Sum"
        sum_ws["C1"] = "Average"
        sum_ws["D1"] = "Count"
        sum_ws["E1"] = "Min"
        sum_ws["F1"] = "Max"
        for cell in sum_ws["A1:F1"]:
            for c in cell:
                c.font = Font(bold=True)
                c.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                c.font = Font(bold=True, color="FFFFFF")
        data = []
        for row in src_ws.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        if data:
            num_cols = len(data[0])
            for col in range(1, num_cols + 1):
                col_letter = get_column_letter(col)
                sum_ws.cell(row=col+1, column=1, value=col_letter)
                sum_ws.cell(row=col+1, column=2, value=f"=SUM('{source_sheet}'!{col_letter}:{col_letter})")
                sum_ws.cell(row=col+1, column=3, value=f"=AVERAGE('{source_sheet}'!{col_letter}:{col_letter})")
                sum_ws.cell(row=col+1, column=4, value=f"=COUNT('{source_sheet}'!{col_letter}:{col_letter})")
                sum_ws.cell(row=col+1, column=5, value=f"=MIN('{source_sheet}'!{col_letter}:{col_letter})")
                sum_ws.cell(row=col+1, column=6, value=f"=MAX('{source_sheet}'!{col_letter}:{col_letter})")
        for col in range(1, 7):
            sum_ws.column_dimensions[get_column_letter(col)].width = 12
        return {"status": "ok", "message": f"Created summary sheet '{summary_name}'"}

    def duplicate_row(self, row: int, times: int = 1, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        row_data = []
        for col in range(1, ws.max_column + 1):
            row_data.append(ws.cell(row=row, column=col).value)
        for i in range(times):
            insert_row = row + i + 1
            ws.insert_rows(insert_row)
            for col, val in enumerate(row_data, 1):
                ws.cell(row=insert_row, column=col, value=val)
        return {"status": "ok", "message": f"Duplicated row {row} {times} time(s)"}

    def delete_empty_rows(self, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        deleted = 0
        rows_to_delete = []
        for row in range(2, ws.max_row + 1):
            is_empty = all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1))
            if is_empty:
                rows_to_delete.append(row)
        for r in reversed(rows_to_delete):
            ws.delete_rows(r)
            deleted += 1
        return {"status": "ok", "message": f"Deleted {deleted} empty rows"}

    def find_and_replace(self, find_text: str, replace_text: str, sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        replaced = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == find_text:
                    cell.value = replace_text
                    replaced += 1
        return {"status": "ok", "message": f"Replaced {replaced} cell(s)"}

    def add_page_setup(self, orientation: str = "portrait", title: str = "Sheet", sheet: Optional[str] = None) -> dict:
        ws = self._get_sheet(sheet)
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = orientation
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.headerFooter.oddHeader.center.text = title
        return {"status": "ok", "message": f"Added page setup ({orientation})"}

    def board_write(self, key: str, value: str) -> dict:
        SharedBlackboard.write(key, value)
        return {"status": "ok", "message": f"Board: wrote '{key}' = '{value}'"}

    def board_read(self, key: str) -> dict:
        val = SharedBlackboard.read(key, "(not set)")
        return {"status": "ok", "message": f"Board: '{key}' = {val}", "value": val}


# ─────────────────────────────────────────────────────────────────────────────
# RAG ANALYZER  (Excel file understanding & context)
# ─────────────────────────────────────────────────────────────────────────────

class RAGAnalyzer:
    def __init__(self):
        self.current_file: Optional[str] = None
        self.schema: dict = {}
        self.data_summary: str = ""
        self.row_count: int = 0
        self.col_count: int = 0

    def analyze_file(self, filepath: str) -> dict:
        if not os.path.exists(filepath):
            return {"status": "error", "message": f"File not found: {filepath}"}
        
        try:
            import pandas as pd
            df = pd.read_excel(filepath, sheet_name=None)
            
            self.current_file = filepath
            self.schema = {}
            all_data = []
            
            for sheet_name, sheet_df in df.items():
                headers = list(sheet_df.columns)
                dtypes = {str(col): str(sheet_df[col].dtype) for col in headers}
                sample_data = sheet_df.head(5).values.tolist()
                
                self.schema[sheet_name] = {
                    "headers": headers,
                    "dtypes": dtypes,
                    "rows": len(sheet_df),
                    "sample": sample_data
                }
                all_data.append({"sheet": sheet_name, "rows": len(sheet_df), "cols": len(headers)})
            
            self.row_count = sum(s["rows"] for s in self.schema.values())
            self.col_count = max((len(s["headers"]) for s in self.schema.values()), default=0)
            
            summary = self._generate_summary()
            
            return {
                "status": "ok",
                "message": f"Analyzed: {os.path.basename(filepath)}",
                "schema": self.schema,
                "summary": summary,
                "file": filepath,
                "sheets": list(self.schema.keys()),
                "stats": all_data
            }
        except Exception as e:
            return {"status": "error", "message": f"Analysis failed: {str(e)}"}

    def _generate_summary(self) -> str:
        lines = [f"File: {os.path.basename(self.current_file)}"]
        lines.append(f"Total Sheets: {len(self.schema)}")
        lines.append(f"Total Rows: {self.row_count}")
        lines.append(f"Total Columns: {self.col_count}")
        lines.append("")
        
        for sheet_name, info in self.schema.items():
            lines.append(f"Sheet '{sheet_name}':")
            lines.append(f"  - Rows: {info['rows']}, Columns: {len(info['headers'])}")
            lines.append(f"  - Columns: {', '.join(info['headers'][:5])}")
            if len(info['headers']) > 5:
                lines.append(f"    ... and {len(info['headers']) - 5} more")
            lines.append("")
        
        return "\n".join(lines)

    def get_context_prompt(self) -> str:
        if not self.current_file:
            return ""
        
        summary = self.data_summary or self._generate_summary()
        context = f"""CURRENT FILE ANALYSIS:
{summary}

You are working with this existing Excel file. The user will give you instructions on what to do with it.
You can read, modify, analyze, or transform the data as requested.
Always save any changes made to this file.

"""
        return context

    def clear(self):
        self.current_file = None
        self.schema = {}
        self.data_summary = ""
        self.row_count = 0
        self.col_count = 0


# ─────────────────────────────────────────────────────────────────────────────
# MULTI-LAYER MEMORY SYSTEM
# ─────────────────────────────────────────────────────────────────────────────

class SharedBlackboard:
    """Telepathy channel - all agents share data instantly without LLM calls"""
    _instance = None
    _data: dict = {}
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._data = {}
        return cls._instance
    
    @classmethod
    def write(cls, key: str, value: Any):
        cls._data[key] = value
    
    @classmethod
    def read(cls, key: str, default: Any = None) -> Any:
        return cls._data.get(key, default)
    
    @classmethod
    def knows(cls, key: str) -> bool:
        return key in cls._data
    
    @classmethod
    def clear(cls):
        cls._data.clear()
    
    @classmethod
    def all_keys(cls) -> list:
        return list(cls._data.keys())
    
    @classmethod
    def dump(cls) -> dict:
        return cls._data.copy()


class WorkingMemory:
    """Keeps last 10 messages, summarizes old half when full"""
    def __init__(self, window_size: int = 10, model_config: dict = None):
        self.messages: list = []
        self.summary: str = ""
        self.window_size = window_size
        self.model_config = model_config or {}
        self._pending_summary = False
    
    def add(self, role: str, content: str):
        self.messages.append({"role": role, "content": content})
        if len(self.messages) > self.window_size:
            self._trigger_summary()
    
    def _trigger_summary(self):
        if self._pending_summary:
            return
        self._pending_summary = True
    
    async def summarize_old_half(self, get_llm_response):
        if len(self.messages) <= self.window_size // 2:
            return
        
        half = len(self.messages) // 2
        old_messages = self.messages[:half]
        summary_prompt = f"Summarize this conversation in 3-5 sentences focusing on key actions and results:\n\n"
        for m in old_messages:
            summary_prompt += f"{m['role']}: {m['content'][:200]}\n"
        
        try:
            if get_llm_response:
                summary = await get_llm_response(summary_prompt)
                self.summary = f"[EARLIER: {summary}]"
                self.messages = self.messages[half:]
                self._pending_summary = False
        except:
            self._pending_summary = False
    
    def get_context(self) -> list:
        ctx = []
        if self.summary:
            ctx.append({"role": "system", "content": self.summary})
        return ctx + self.messages[-self.window_size:]


class SemanticMemory:
    """Vector store - embeds messages, retrieves top-k relevant past facts"""
    def __init__(self, embedding_model: str = "nomic-embed-text"):
        self.embeddings: list = []
        self.embedding_model = embedding_model
        self._client = None
    
    def _get_embedding(self, text: str) -> list:
        try:
            import ollama
            if not self._client:
                self._client = ollama
            result = self._client.embeddings(model=self.embedding_model, prompt=text)
            return result.embedding
        except Exception:
            return [0.0] * 768
    
    def _cosine_similarity(self, a: list, b: list) -> float:
        dot = sum(x*y for x, y in zip(a, b))
        norm_a = sum(x*x for x in a) ** 0.5
        norm_b = sum(x*x for x in b) ** 0.5
        if norm_a == 0 or norm_b == 0:
            return 0.0
        return dot / (norm_a * norm_b)
    
    def add_message(self, text: str, metadata: dict = None):
        embedding = self._get_embedding(text)
        self.embeddings.append({
            "text": text,
            "embedding": embedding,
            "metadata": metadata or {}
        })
    
    def retrieve(self, query: str, top_k: int = 4) -> list:
        if not self.embeddings:
            return []
        
        query_embedding = self._get_embedding(query)
        similarities = []
        
        for i, entry in enumerate(self.embeddings):
            sim = self._cosine_similarity(query_embedding, entry["embedding"])
            similarities.append((i, sim))
        
        similarities.sort(key=lambda x: x[1], reverse=True)
        results = []
        for i, sim in similarities[:top_k]:
            results.append({
                "text": self.embeddings[i]["text"],
                "relevance": sim,
                "metadata": self.embeddings[i]["metadata"]
            })
        return results
    
    def get_retrieval_context(self, query: str, top_k: int = 4) -> str:
        results = self.retrieve(query, top_k)
        if not results:
            return ""
        
        context = "RELEVANT PAST FACTS:\n"
        for r in results:
            context += f"- {r['text'][:150]}\n"
        return context
    
    def clear(self):
        self.embeddings = []


class EpisodicMemory:
    """Session summaries - persisted to disk, loaded on restart"""
    def __init__(self, storage_path: str = "episodes.json"):
        self.storage_path = storage_path
        self.episodes: list = []
        self.current_episode: list = []
        self._load_episodes()
    
    def _load_episodes(self):
        if os.path.exists(self.storage_path):
            try:
                with open(self.storage_path, 'r') as f:
                    self.episodes = json.load(f)
            except:
                self.episodes = []
    
    def _save_episodes(self):
        with open(self.storage_path, 'w') as f:
            json.dump(self.episodes, f, indent=2)
    
    def add_turn(self, role: str, content: str, is_tool_result: bool = False):
        self.current_episode.append({"role": role, "content": content, "is_tool": is_tool_result})
    
    async def end_session(self, get_llm_response):
        if not self.current_episode:
            return
        
        summary_prompt = "Compress this entire conversation into 4-6 sentences summarizing what was done:\n\n"
        for turn in self.current_episode:
            if not turn.get("is_tool"):
                summary_prompt += f"{turn['role']}: {turn['content'][:150]}\n"
        
        try:
            if get_llm_response:
                summary = await get_llm_response(summary_prompt)
                self.episodes.append({
                    "summary": summary,
                    "timestamp": datetime.now().isoformat(),
                    "turn_count": len(self.current_episode)
                })
                if len(self.episodes) > 20:
                    self.episodes = self.episodes[-20:]
                self._save_episodes()
        except:
            pass
        
        self.current_episode = []
    
    def get_relevant_episodes(self, query: str, top_k: int = 3) -> list:
        if not self.episodes:
            return []
        
        query_lower = query.lower()
        scored = []
        for ep in self.episodes:
            score = sum(1 for word in ep["summary"].lower().split() if word in query_lower)
            scored.append((ep, score))
        
        scored.sort(key=lambda x: x[1], reverse=True)
        return [ep["summary"] for ep, _ in scored[:top_k]]
    
    def get_episodic_context(self, query: str) -> str:
        relevant = self.get_relevant_episodes(query)
        if not relevant:
            return ""
        
        context = "PAST SESSIONS:\n"
        for i, summary in enumerate(relevant, 1):
            context += f"{i}. {summary}\n"
        return context
    
    def clear(self):
        self.episodes = []
        self._save_episodes()


class MemoryManager:
    """Unified memory interface - combines all memory layers"""
    def __init__(self, model_config: dict = None):
        self.working = WorkingMemory(10, model_config)
        self.semantic = SemanticMemory()
        self.episodic = EpisodicMemory()
        self.blackboard = SharedBlackboard()
        self._session_active = False
    
    def start_session(self):
        self._session_active = True
    
    async def end_session(self, get_llm_response=None):
        self._session_active = False
        await self.episodic.end_session(get_llm_response)
    
    def add_message(self, role: str, content: str, is_tool_result: bool = False):
        if self._session_active:
            self.working.add(role, content)
            self.semantic.add_message(content, {"role": role})
            self.episodic.add_turn(role, content, is_tool_result)
    
    def get_full_context(self, current_prompt: str) -> list:
        ctx = []
        
        episodic = self.episodic.get_episodic_context(current_prompt)
        if episodic:
            ctx.append({"role": "system", "content": episodic})
        
        semantic = self.semantic.get_retrieval_context(current_prompt, 4)
        if semantic:
            ctx.append({"role": "system", "content": semantic})
        
        working_ctx = self.working.get_context()
        ctx.extend(working_ctx)
        
        return ctx
    
    def board_write(self, key: str, value: Any):
        SharedBlackboard.write(key, value)
    
    def board_read(self, key: str, default: Any = None) -> Any:
        return SharedBlackboard.read(key, default)
    
    def board_knows(self, key: str) -> bool:
        return SharedBlackboard.knows(key)
    
    async def summarize_working_memory(self, get_llm_response):
        await self.working.summarize_old_half(get_llm_response)
    
    def clear_all(self):
        self.working.messages = []
        self.working.summary = ""
        self.semantic.clear()
        self.episodic.clear()
        SharedBlackboard.clear()


# ─────────────────────────────────────────────────────────────────────────────
# TOOL REGISTRY  (JSON schema for LLM function calling)
# ─────────────────────────────────────────────────────────────────────────────

TOOL_SCHEMAS = [
    {
        "type": "function",
        "function": {
            "name": "create_workbook",
            "description": "Create a new Excel workbook and save it to a file path",
            "parameters": {
                "type": "object",
                "properties": {"filepath": {"type": "string", "description": "Full path to save the .xlsx file"}},
                "required": ["filepath"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "open_workbook",
            "description": "Open an existing Excel workbook from a file path",
            "parameters": {
                "type": "object",
                "properties": {"filepath": {"type": "string"}},
                "required": ["filepath"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "save_workbook",
            "description": "Save the current workbook to disk",
            "parameters": {
                "type": "object",
                "properties": {"filepath": {"type": "string", "description": "Optional path override"}}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_sheet",
            "description": "Create a new worksheet in the current workbook",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "position": {"type": "integer", "description": "0-indexed position"}
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "set_active_sheet",
            "description": "Switch to a different worksheet",
            "parameters": {
                "type": "object",
                "properties": {"name": {"type": "string"}},
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "rename_sheet",
            "description": "Rename a worksheet",
            "parameters": {
                "type": "object",
                "properties": {
                    "old_name": {"type": "string"},
                    "new_name": {"type": "string"}
                },
                "required": ["old_name", "new_name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_sheet",
            "description": "Delete a worksheet",
            "parameters": {
                "type": "object",
                "properties": {"name": {"type": "string"}},
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "copy_sheet",
            "description": "Copy a worksheet to a new sheet",
            "parameters": {
                "type": "object",
                "properties": {
                    "source": {"type": "string"},
                    "target": {"type": "string"}
                },
                "required": ["source", "target"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_cell",
            "description": "Write a single value or formula to a cell",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {"type": "string", "description": "Cell reference like A1"},
                    "value": {"description": "Value to write (string, number, etc.)"},
                    "sheet": {"type": "string", "description": "Optional sheet name"}
                },
                "required": ["cell", "value"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_range",
            "description": "Write a 2D array of data starting from a cell",
            "parameters": {
                "type": "object",
                "properties": {
                    "start_cell": {"type": "string"},
                    "data": {"type": "array", "items": {"type": "array"}, "description": "2D array of values"},
                    "sheet": {"type": "string"}
                },
                "required": ["start_cell", "data"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_cell",
            "description": "Read the value of a single cell",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {"type": "string"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_range",
            "description": "Read a range of cells as a 2D array",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string", "description": "Range like A1:D10"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "set_formula",
            "description": "Set an Excel formula in a cell",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {"type": "string"},
                    "formula": {"type": "string", "description": "Formula with or without = prefix"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell", "formula"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "format_cells",
            "description": "Apply formatting to a range of cells",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string"},
                    "options": {
                        "type": "object",
                        "description": "Formatting options",
                        "properties": {
                            "font_bold": {"type": "boolean"},
                            "font_size": {"type": "number"},
                            "font_color": {"type": "string", "description": "Hex color like FF0000"},
                            "font_name": {"type": "string"},
                            "font_italic": {"type": "boolean"},
                            "bg_color": {"type": "string", "description": "Hex color like #FF0000"},
                            "alignment": {"type": "object"},
                            "border": {"type": "object"},
                            "number_format": {"type": "string"}
                        }
                    },
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range", "options"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "set_column_width",
            "description": "Set the width of a column",
            "parameters": {
                "type": "object",
                "properties": {
                    "col": {"type": "string", "description": "Column letter like A"},
                    "width": {"type": "number"},
                    "sheet": {"type": "string"}
                },
                "required": ["col", "width"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "merge_cells",
            "description": "Merge a range of cells",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_chart",
            "description": "Add a chart to the worksheet",
            "parameters": {
                "type": "object",
                "properties": {
                    "chart_type": {"type": "string", "enum": ["bar", "line", "scatter", "Line"]},
                    "data_range": {"type": "string", "description": "Cell range for chart data (e.g., A1:D10)"},
                    "title": {"type": "string"},
                    "position": {"type": "string", "description": "Cell position for chart, e.g. E1"},
                    "sheet": {"type": "string"},
                    "chart_data": {"type": "object", "description": "Alternative: dict with 'headers' and 'rows' keys (AI-generated format)"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_pivot_summary",
            "description": "Create a pivot-style summary table from data",
            "parameters": {
                "type": "object",
                "properties": {
                    "source_range": {"type": "string"},
                    "target_cell": {"type": "string"},
                    "row_field": {"type": "integer", "description": "0-indexed column for row labels"},
                    "value_field": {"type": "integer", "description": "0-indexed column for values to sum"},
                    "sheet": {"type": "string"},
                    "target_sheet": {"type": "string"}
                },
                "required": ["source_range", "target_cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_conditional_formatting",
            "description": "Add conditional formatting rules to a range",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string"},
                    "rule_type": {"type": "string", "enum": ["color_scale", "cell_is"]},
                    "options": {"type": "object"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range", "rule_type", "options"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_autofilter",
            "description": "Add AutoFilter to a range",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "freeze_panes",
            "description": "Freeze rows/columns above and to the left of a cell",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell": {"type": "string"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "insert_row",
            "description": "Insert blank rows",
            "parameters": {
                "type": "object",
                "properties": {
                    "row": {"type": "integer"},
                    "count": {"type": "integer"},
                    "sheet": {"type": "string"}
                },
                "required": ["row"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_row",
            "description": "Delete rows",
            "parameters": {
                "type": "object",
                "properties": {
                    "row": {"type": "integer"},
                    "count": {"type": "integer"},
                    "sheet": {"type": "string"}
                },
                "required": ["row"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_workbook_info",
            "description": "Get metadata about the current workbook",
            "parameters": {"type": "object", "properties": {}}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "sort_range",
            "description": "Sort a range of data by a specific column (ascending or descending)",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string", "description": "Range to sort like A1:D100"},
                    "key_column": {"type": "integer", "description": "0-indexed column number to sort by"},
                    "ascending": {"type": "boolean", "description": "True for ascending, False for descending"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range", "key_column"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_totals_row",
            "description": "Add a totals row with SUM formulas to calculate column sums",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string", "description": "Range including headers and data"},
                    "label": {"type": "string", "description": "Label for the totals row (default: 'Total')"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "apply_header_style",
            "description": "Apply professional header formatting (bold, white text, dark background)",
            "parameters": {
                "type": "object",
                "properties": {
                    "row": {"type": "integer", "description": "Header row number (default: 1)"},
                    "sheet": {"type": "string"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "apply_alternating_rows",
            "description": "Apply alternating row colors for better readability",
            "parameters": {
                "type": "object",
                "properties": {
                    "start_row": {"type": "integer", "description": "First data row (default: 2)"},
                    "sheet": {"type": "string"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "auto_size_columns",
            "description": "Automatically set column widths to optimal size",
            "parameters": {
                "type": "object",
                "properties": {"sheet": {"type": "string"}}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_data_bars",
            "description": "Add conditional formatting data bars to visualize values",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string", "description": "Range to apply data bars"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "highlight_top_values",
            "description": "Highlight top N values with conditional formatting",
            "parameters": {
                "type": "object",
                "properties": {
                    "cell_range": {"type": "string"},
                    "n": {"type": "integer", "description": "Number of top values to highlight"},
                    "sheet": {"type": "string"}
                },
                "required": ["cell_range", "n"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_status_column",
            "description": "Add a new column with IF formula to classify values as Good/Review based on threshold",
            "parameters": {
                "type": "object",
                "properties": {
                    "threshold_col": {"type": "integer", "description": "Column index (1-based) to check"},
                    "threshold_value": {"type": "number", "description": "Threshold value"},
                    "new_col_name": {"type": "string", "description": "Name for new status column"},
                    "sheet": {"type": "string"}
                },
                "required": ["threshold_col", "threshold_value"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_summary_sheet",
            "description": "Create a new sheet with statistical summary (Sum, Average, Count, Min, Max) for each column",
            "parameters": {
                "type": "object",
                "properties": {
                    "source_sheet": {"type": "string", "description": "Sheet to analyze"},
                    "summary_name": {"type": "string", "description": "Name for summary sheet (default: Summary)"},
                    "sheet": {"type": "string"}
                },
                "required": ["source_sheet"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "duplicate_row",
            "description": "Duplicate a specific row multiple times",
            "parameters": {
                "type": "object",
                "properties": {
                    "row": {"type": "integer", "description": "Row number to duplicate"},
                    "times": {"type": "integer", "description": "How many times to duplicate"},
                    "sheet": {"type": "string"}
                },
                "required": ["row", "times"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_empty_rows",
            "description": "Remove empty rows from the sheet",
            "parameters": {
                "type": "object",
                "properties": {"sheet": {"type": "string"}}
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_and_replace",
            "description": "Find text and replace with new value throughout the sheet",
            "parameters": {
                "type": "object",
                "properties": {
                    "find_text": {"type": "string"},
                    "replace_text": {"type": "string"},
                    "sheet": {"type": "string"}
                },
                "required": ["find_text", "replace_text"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_page_setup",
            "description": "Configure page setup (orientation, fit to page, headers)",
            "parameters": {
                "type": "object",
                "properties": {
                    "orientation": {"type": "string", "enum": ["portrait", "landscape"]},
                    "title": {"type": "string", "description": "Sheet title for print header"},
                    "sheet": {"type": "string"}
                }
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "board_write",
            "description": "Write a key-value pair to the shared blackboard (telepathy channel for agent coordination)",
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {"type": "string", "description": "Key name to store"},
                    "value": {"type": "string", "description": "Value to store"}
                },
                "required": ["key", "value"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "board_read",
            "description": "Read a value from the shared blackboard by key",
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {"type": "string", "description": "Key to read"}
                },
                "required": ["key"]
            }
        }
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# AI AGENT  (Ollama / Groq dual-routing)
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are SynthGrid's Excel Automation Agent. You understand natural language and execute exactly what the user asks.

MEMORY SYSTEMS: You have access to multiple memory layers that help you:
- WORKING MEMORY: You see recent conversation context (last ~10 turns)
- SEMANTIC MEMORY: Relevant past facts are retrieved based on similarity
- EPISODIC MEMORY: Past session summaries are available for context
- SHARED BLACKBOARD: Use board_write(key, value) to store key info, board_read(key) to retrieve

CORE INSTRUCTIONS:
1. UNDERSTAND INTENT: Parse what the user WANTS, not just what they say. If they say "make it pretty", apply professional formatting. If they say "analyze this", explore the data and create summaries.

2. EXECUTE EXACTLY: Do exactly what is asked. If user says "add totals row", add a row with SUM formulas. If they say "sort by revenue", sort the data. DO NOT ask clarification unless absolutely necessary.

3. BE PROACTIVE: If user asks to "analyze", automatically create pivot tables, charts, insights. If they ask to "visualize", create appropriate charts. Think ahead about what they likely want.

4. WORK WITH EXISTING DATA: When a file is loaded via RAG, analyze its structure first. Understand column types (text, numbers, dates) and create appropriate transformations.

5. COMPLETE WORKFLOW: For any task, do ALL logical steps:
   - Creating data → format headers, add formulas, add charts
   - Analyzing → create summaries, pivot tables, visualizations
   - Visualizing → choose appropriate chart type for data
   - Formatting → apply consistent, professional styling

6. TOOL USAGE: You have tools for EVERY Excel operation. Use them freely:
   - Read/Write cells, ranges, formulas
   - Create sheets, rename, delete, copy
   - Format cells (colors, fonts, borders, alignment)
   - Add charts (bar, line, scatter, pie)
   - Create pivot summaries
   - Add data validation, filters, freeze panes
   - Conditional formatting
   - board_write(key, value) - store info for other agents
   - board_read(key) - retrieve stored info

7. NEVER DESCRIBE - ALWAYS DO: If asked to "add a chart", create it. If asked to "format headers", format them. Execute, don't explain what you would do.

8. SAVE ALWAYS: After any modifications, save the workbook.

EXAMPLES OF EXECUTING EXACTLY:
- "Add a totals column" → Write SUM formulas in new column
- "Make row 1 bold with dark background" → Apply formatting to row 1
- "Show monthly trends" → Create line chart with monthly data
- "Find top 10 values" → Sort and highlight top 10
- "Add conditional formatting" → Apply color scales based on values
- "Create a dashboard" → Combine summary, charts, key metrics
- "Store this for later" → Use board_write("key", "value")
- "Get that info" → Use board_read("key")

Remember: The user wants you to DO what they say. Execute completely, professionally, and save results."""


class AgentWorker(QThread):
    """Background thread for AI inference + tool execution"""
    log_signal = pyqtSignal(str, str)   # (message, level)
    grid_signal = pyqtSignal()
    done_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

    def __init__(self, prompt: str, engine: ExcelEngine, model_config: dict, rag_context: str = "", memory: MemoryManager = None):
        super().__init__()
        self.prompt = prompt
        self.engine = engine
        self.model_config = model_config
        self.rag_context = rag_context
        self.memory = memory
        self._messages = []

    def log(self, msg: str, level: str = "info"):
        self.log_signal.emit(msg, level)

    def execute_tool(self, name: str, args: dict) -> str:
        method = getattr(self.engine, name, None)
        if not method:
            return json.dumps({"status": "error", "message": f"Unknown tool: {name}"})
        try:
            result = method(**args)
            self.grid_signal.emit()
            return json.dumps(result)
        except Exception as e:
            return json.dumps({"status": "error", "message": str(e), "traceback": traceback.format_exc()})

    def run_with_groq(self, messages):
        try:
            from groq import Groq
            api_key = self.model_config.get("groq_api_key", os.environ.get("GROQ_API_KEY", ""))
            if not api_key:
                raise ValueError("GROQ_API_KEY not set")
            client = Groq(api_key=api_key)
            model = self.model_config.get("groq_model", "llama-3.3-70b-versatile")
            self.log(f"◈ GROQ REQUEST → model={model}", "api")
            t0 = time.time()
            resp = client.chat.completions.create(
                model=model,
                messages=messages,
                tools=TOOL_SCHEMAS,
                tool_choice="auto",
                max_tokens=4096,
            )
            elapsed = time.time() - t0
            usage = resp.usage
            tps = (usage.completion_tokens / elapsed) if elapsed > 0 else 0
            self.log(f"◈ GROQ RESPONSE ← {usage.completion_tokens} tokens | {tps:.1f} tok/s | {elapsed:.2f}s", "api")
            return resp.choices[0].message
        except Exception as e:
            raise RuntimeError(f"Groq error: {e}")

    def run_with_ollama(self, messages):
        try:
            import ollama
            model = self.model_config.get("ollama_model", "llama3.1")
            self.log(f"◈ OLLAMA REQUEST → model={model}", "api")
            t0 = time.time()
            
            # Convert tool schemas to ollama format
            tools = []
            for t in TOOL_SCHEMAS:
                tools.append({
                    "type": "function",
                    "function": t["function"]
                })
            
            resp = ollama.chat(
                model=model,
                messages=messages,
                tools=tools,
                stream=False
            )
            elapsed = time.time() - t0
            self.log(f"◈ OLLAMA RESPONSE ← {elapsed:.2f}s", "api")
            return resp.message
        except Exception as e:
            raise RuntimeError(f"Ollama error: {e}")

    def run_with_openai(self, messages):
        try:
            from openai import OpenAI
            api_key = self.model_config.get("openai_api_key", os.environ.get("OPENAI_API_KEY", ""))
            if not api_key:
                raise ValueError("OPENAI_API_KEY not set")
            client = OpenAI(api_key=api_key)
            model = self.model_config.get("openai_model", "gpt-4o")
            self.log(f"◈ OPENAI REQUEST → model={model}", "api")
            t0 = time.time()
            resp = client.chat.completions.create(
                model=model, messages=messages, tools=TOOL_SCHEMAS,
                tool_choice="auto", max_tokens=4096,
            )
            elapsed = time.time() - t0
            u = resp.usage
            tps = (u.completion_tokens / elapsed) if elapsed > 0 else 0
            self.log(f"◈ OPENAI RESPONSE ← {u.completion_tokens} tokens | {tps:.1f} tok/s | {elapsed:.2f}s", "api")
            return resp.choices[0].message
        except Exception as e:
            raise RuntimeError(f"OpenAI error: {e}")

    def run_with_anthropic(self, messages):
        try:
            from anthropic import Anthropic
            api_key = self.model_config.get("anthropic_api_key", os.environ.get("ANTHROPIC_API_KEY", ""))
            if not api_key:
                raise ValueError("ANTHROPIC_API_KEY not set")
            client = Anthropic(api_key=api_key)
            model = self.model_config.get("anthropic_model", "claude-3-5-sonnet-20241022")
            self.log(f"◈ ANTHROPIC REQUEST → model={model}", "api")
            t0 = time.time()

            ant_tools = [{"name": t["function"]["name"], "description": t["function"].get("description", ""), "input_schema": t["function"]["parameters"]} for t in TOOL_SCHEMAS]
            system = next((m["content"] for m in messages if m["role"] == "system"), "")

            ant_messages = []
            for m in messages:
                r = m["role"]
                if r == "system":
                    continue
                if r == "tool":
                    ant_messages.append({"role": "user", "content": [{"type": "tool_result", "tool_use_id": m["tool_call_id"], "content": m["content"]}]})
                elif r == "assistant" and "tool_calls" in m:
                    c = []
                    if m.get("content"):
                        c.append({"type": "text", "text": m["content"]})
                    for tc in m["tool_calls"]:
                        c.append({"type": "tool_use", "id": tc["id"], "name": tc["function"]["name"], "input": tc["function"]["arguments"]})
                    ant_messages.append({"role": "assistant", "content": c})
                else:
                    ant_messages.append({"role": r, "content": m.get("content", "")})

            resp = client.messages.create(model=model, system=system, messages=ant_messages, tools=ant_tools or None, max_tokens=4096)
            elapsed = time.time() - t0

            # Normalize Anthropic response to match OpenAI message interface
            content = ""
            tool_calls = []
            for block in resp.content:
                if block.type == "text":
                    content = block.text
                elif block.type == "tool_use":
                    tool_calls.append(type('ToolCall', (), {'id': block.id, 'function': type('Fn', (), {'name': block.name, 'arguments': json.dumps(block.input)})})())

            msg = type('Msg', (), {'content': content, 'tool_calls': tool_calls or None})()
            self.log(f"◈ ANTHROPIC RESPONSE ← {elapsed:.2f}s", "api")
            return msg
        except Exception as e:
            raise RuntimeError(f"Anthropic error: {e}")

    def run(self):
        try:
            user_content = self.prompt
            if self.rag_context:
                user_content = self.rag_context + "\n\nUSER REQUEST: " + self.prompt
                self.log(f"◈ RAG CONTEXT INJECTED", "info")
            
            base_messages = [{"role": "system", "content": SYSTEM_PROMPT}]
            
            if self.memory:
                self.memory.start_session()
                self.memory.add_message("user", self.prompt)
                full_context = self.memory.get_full_context(self.prompt)
                for msg in full_context:
                    if msg not in base_messages:
                        base_messages.append(msg)
                self.log(f"◈ MEMORY CONTEXT: {len(full_context)} messages loaded", "info")
            
            self._messages = base_messages + [{"role": "user", "content": user_content}]
            self.log(f"▶ AGENT START: {self.prompt[:80]}...", "start")
            
            backend = self.model_config.get("backend", "ollama")
            max_iterations = 10
            
            for iteration in range(max_iterations):
                self.log(f"── ITERATION {iteration + 1} ──────────────────", "divider")
                
                if backend == "groq":
                    message = self.run_with_groq(self._messages)
                elif backend == "openai":
                    message = self.run_with_openai(self._messages)
                elif backend == "anthropic":
                    message = self.run_with_anthropic(self._messages)
                else:
                    message = self.run_with_ollama(self._messages)

                # Check for tool calls
                tool_calls = getattr(message, "tool_calls", None)
                
                if not tool_calls:
                    # Final text response
                    content = getattr(message, "content", "") or ""
                    self.log(f"✓ AGENT FINAL: {content}", "success")
                    self.done_signal.emit(content)
                    return

                # Process tool calls
                msg_dict = {"role": "assistant", "content": getattr(message, "content", "") or ""}
                if tool_calls:
                    msg_dict["tool_calls"] = [
                        {
                            "id": tc.id if hasattr(tc, 'id') else f"call_{i}",
                            "type": "function",
                            "function": {
                                "name": tc.function.name,
                                "arguments": json.loads(tc.function.arguments) if isinstance(tc.function.arguments, str) else tc.function.arguments
                            }
                        }
                        for i, tc in enumerate(tool_calls)
                    ]
                self._messages.append(msg_dict)

                for tc in tool_calls:
                    fn_name = tc.function.name
                    fn_args_raw = tc.function.arguments
                    tc_id = tc.id if hasattr(tc, 'id') else f"call_0"
                    
                    if isinstance(fn_args_raw, str):
                        try:
                            fn_args = json.loads(fn_args_raw)
                        except:
                            fn_args = {}
                    else:
                        fn_args = fn_args_raw or {}

                    self.log(f"⚡ TOOL CALL → {fn_name}({json.dumps(fn_args)[:120]})", "tool")
                    result_str = self.execute_tool(fn_name, fn_args)
                    result_obj = json.loads(result_str)
                    
                    status = result_obj.get("status", "?")
                    if status == "ok":
                        msg_part = result_obj.get("message", result_str[:100])
                        self.log(f"  ✓ RESULT: {msg_part}", "ok")
                    else:
                        self.log(f"  ✗ ERROR: {result_obj.get('message', result_str[:100])}", "error")

                    self._messages.append({
                        "role": "tool",
                        "tool_call_id": tc_id,
                        "content": result_str
                    })

            self.log("⚠ Max iterations reached", "warn")
            self.done_signal.emit("TASK PAUSED - Say 'continue' to resume")
        except Exception as e:
            self.error_signal.emit(str(e))
            self.log(f"✗ FATAL: {e}", "error")


# ─────────────────────────────────────────────────────────────────────────────
# STYLING  — Cybersigilism Dark Industrial
# ─────────────────────────────────────────────────────────────────────────────

PALETTE = {
    "bg0":      "#050508",
    "bg1":      "#0a0a10",
    "bg2":      "#0f0f18",
    "bg3":      "#14141f",
    "panel":    "#0c0c16",
    "border":   "#1a1a2e",
    "border2":  "#252540",
    "accent":   "#00ff9f",
    "accent2":  "#00ccff",
    "accent3":  "#ff3366",
    "warn":     "#ffaa00",
    "text":     "#c8d8e8",
    "text2":    "#7888a0",
    "text3":    "#405060",
    "green":    "#39ff14",
    "red":      "#ff2244",
    "cyan":     "#00e5ff",
    "magenta":  "#ff00cc",
    "yellow":   "#ffe000",
}

GLOBAL_STYLE = f"""
QMainWindow, QWidget {{
    background-color: {PALETTE['bg0']};
    color: {PALETTE['text']};
    font-family: 'JetBrains Mono', 'Fira Code', 'Courier New', monospace;
    font-size: 12px;
}}
QFrame#panel {{
    background-color: {PALETTE['panel']};
    border: 1px solid {PALETTE['border']};
}}
QLabel#header {{
    color: {PALETTE['accent']};
    font-size: 10px;
    font-weight: bold;
    letter-spacing: 3px;
    padding: 2px 4px;
    background-color: {PALETTE['bg1']};
    border-bottom: 1px solid {PALETTE['border2']};
}}
QTextEdit {{
    background-color: {PALETTE['bg1']};
    color: {PALETTE['text']};
    border: 1px solid {PALETTE['border']};
    selection-background-color: {PALETTE['accent']}44;
    font-family: 'JetBrains Mono', 'Fira Code', monospace;
    font-size: 11px;
    padding: 4px;
}}
QLineEdit {{
    background-color: {PALETTE['bg2']};
    color: {PALETTE['accent']};
    border: 1px solid {PALETTE['accent']}66;
    border-bottom: 2px solid {PALETTE['accent']};
    padding: 6px 10px;
    font-family: 'JetBrains Mono', 'Fira Code', monospace;
    font-size: 13px;
    selection-background-color: {PALETTE['accent']}44;
}}
QLineEdit:focus {{
    border-color: {PALETTE['accent']};
    background-color: {PALETTE['bg3']};
}}
QPushButton {{
    background-color: {PALETTE['bg2']};
    color: {PALETTE['accent']};
    border: 1px solid {PALETTE['accent']}66;
    padding: 5px 14px;
    font-family: 'JetBrains Mono', monospace;
    font-size: 11px;
    font-weight: bold;
    letter-spacing: 1px;
}}
QPushButton:hover {{
    background-color: {PALETTE['accent']}22;
    border-color: {PALETTE['accent']};
    color: {PALETTE['bg0']};
    background-color: {PALETTE['accent']};
}}
QPushButton:pressed {{
    background-color: {PALETTE['accent']}44;
}}
QPushButton#danger {{
    color: {PALETTE['red']};
    border-color: {PALETTE['red']}66;
}}
QPushButton#danger:hover {{
    background-color: {PALETTE['red']};
    color: {PALETTE['bg0']};
}}
QComboBox {{
    background-color: {PALETTE['bg2']};
    color: {PALETTE['accent2']};
    border: 1px solid {PALETTE['border2']};
    padding: 4px 8px;
    font-family: monospace;
}}
QComboBox::drop-down {{ border: none; }}
QComboBox QAbstractItemView {{
    background-color: {PALETTE['bg1']};
    color: {PALETTE['text']};
    selection-background-color: {PALETTE['accent']}33;
}}
QTableWidget {{
    background-color: {PALETTE['bg1']};
    color: {PALETTE['text']};
    border: none;
    gridline-color: {PALETTE['border']};
    font-family: monospace;
    font-size: 11px;
}}
QTableWidget::item {{
    padding: 2px 6px;
    border-bottom: 1px solid {PALETTE['border']};
}}
QTableWidget::item:selected {{
    background-color: {PALETTE['accent']}33;
    color: {PALETTE['accent']};
}}
QHeaderView::section {{
    background-color: {PALETTE['bg2']};
    color: {PALETTE['accent2']};
    border: none;
    border-right: 1px solid {PALETTE['border']};
    border-bottom: 1px solid {PALETTE['border2']};
    padding: 3px 6px;
    font-weight: bold;
    font-size: 10px;
    letter-spacing: 1px;
}}
QScrollBar:vertical {{
    background: {PALETTE['bg1']};
    width: 6px;
    border: none;
}}
QScrollBar::handle:vertical {{
    background: {PALETTE['accent']}44;
    min-height: 20px;
}}
QScrollBar::handle:vertical:hover {{
    background: {PALETTE['accent']}88;
}}
QScrollBar:horizontal {{
    background: {PALETTE['bg1']};
    height: 6px;
    border: none;
}}
QScrollBar::handle:horizontal {{
    background: {PALETTE['accent']}44;
}}
QSplitter::handle {{
    background-color: {PALETTE['border2']};
    width: 2px;
    height: 2px;
}}
QTabWidget::pane {{
    background-color: {PALETTE['panel']};
    border: 1px solid {PALETTE['border']};
}}
QTabBar::tab {{
    background-color: {PALETTE['bg1']};
    color: {PALETTE['text2']};
    padding: 5px 14px;
    border: 1px solid {PALETTE['border']};
    font-size: 10px;
    letter-spacing: 1px;
}}
QTabBar::tab:selected {{
    background-color: {PALETTE['bg3']};
    color: {PALETTE['accent']};
    border-bottom: 2px solid {PALETTE['accent']};
}}
QProgressBar {{
    background-color: {PALETTE['bg2']};
    border: 1px solid {PALETTE['border2']};
    height: 4px;
    text-align: center;
    color: transparent;
}}
QProgressBar::chunk {{
    background-color: {PALETTE['accent']};
}}
QStatusBar {{
    background-color: {PALETTE['bg1']};
    color: {PALETTE['text2']};
    border-top: 1px solid {PALETTE['border']};
    font-size: 10px;
}}
QGroupBox {{
    border: 1px solid {PALETTE['border2']};
    margin-top: 8px;
    padding-top: 8px;
    color: {PALETTE['text2']};
    font-size: 10px;
}}
QGroupBox::title {{
    color: {PALETTE['accent2']};
    subcontrol-origin: margin;
    left: 8px;
}}
"""

LOG_COLORS = {
    "info":    PALETTE["text"],
    "api":     PALETTE["cyan"],
    "tool":    PALETTE["yellow"],
    "ok":      PALETTE["green"],
    "error":   PALETTE["red"],
    "warn":    PALETTE["warn"],
    "success": PALETTE["accent"],
    "start":   PALETTE["magenta"],
    "divider": PALETTE["border2"],
}


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM WIDGETS
# ─────────────────────────────────────────────────────────────────────────────

class SigilBorder(QFrame):
    """Frame with decorative corner sigils"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("panel")

    def paintEvent(self, event):
        super().paintEvent(event)
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing, False)
        
        accent = QColor(PALETTE["accent"])
        accent.setAlpha(120)
        border2 = QColor(PALETTE["border2"])
        
        p.setPen(QPen(accent, 1))
        r = self.rect()
        sz = 10

        # Corner decorations
        corners = [
            (r.left(), r.top()),
            (r.right(), r.top()),
            (r.left(), r.bottom()),
            (r.right(), r.bottom()),
        ]
        for cx, cy in corners:
            dx = 1 if cx == r.left() else -1
            dy = 1 if cy == r.top() else -1
            p.drawLine(cx, cy, cx + dx * sz, cy)
            p.drawLine(cx, cy, cx, cy + dy * sz)
            p.drawPoint(cx + dx * sz, cy + dy * 3)
            p.drawPoint(cx + dx * 3, cy + dy * sz)


class SectionHeader(QLabel):
    """Styled section header with sigil decorations"""
    def __init__(self, text: str, parent=None):
        super().__init__(parent)
        self.setObjectName("header")
        self.setText(f"  ◈  {text}  ◈  ")
        self.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.setFixedHeight(24)


class LogConsole(QTextEdit):
    """Colored log output console"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self.document().setMaximumBlockCount(2000)

    def append_log(self, message: str, level: str = "info"):
        color = LOG_COLORS.get(level, PALETTE["text"])
        ts = datetime.now().strftime("%H:%M:%S.%f")[:12]
        
        cursor = self.textCursor()
        cursor.movePosition(cursor.MoveOperation.End)
        
        fmt = QTextCharFormat()
        
        # Timestamp
        fmt.setForeground(QColor(PALETTE["text3"]))
        cursor.insertText(f"[{ts}] ", fmt)
        
        # Message  
        fmt.setForeground(QColor(color))
        if level in ("ok", "success"):
            fmt.setFontWeight(700)
        else:
            fmt.setFontWeight(400)
        cursor.insertText(message + "\n", fmt)
        
        self.setTextCursor(cursor)
        self.ensureCursorVisible()


class GridView(QTableWidget):
    """Excel data preview table"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.verticalHeader().setDefaultSectionSize(22)
        self.setAlternatingRowColors(False)
        
    def load_data(self, data: list):
        if not data:
            self.setRowCount(0)
            self.setColumnCount(0)
            return
        
        max_cols = max(len(row) for row in data) if data else 0
        self.setRowCount(len(data))
        self.setColumnCount(max_cols)
        
        # Column headers
        col_headers = [get_column_letter(i+1) for i in range(max_cols)]
        self.setHorizontalHeaderLabels(col_headers)
        self.setVerticalHeaderLabels([str(i+1) for i in range(len(data))])
        
        for r, row in enumerate(data):
            for c, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val is not None else "")
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                # Color header row
                if r == 0 and val is not None:
                    item.setForeground(QColor(PALETTE["accent"]))
                    item.setBackground(QColor(PALETTE["bg2"]))
                elif val is not None and isinstance(val, (int, float)):
                    item.setForeground(QColor(PALETTE["accent2"]))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.setItem(r, c, item)


class NodeSelector(QWidget):
    """AI model selector and status panel"""
    config_changed = pyqtSignal(dict)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()
        self._load_keys()
        self._config = self._get_config()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(12, 8, 12, 8)

        header = QLabel("🤖 AI Settings")
        header.setStyleSheet(f"color: {PALETTE['accent2']}; font-size: 14px; font-weight: bold;")
        layout.addWidget(header)

        backend_label = QLabel("Select AI Backend:")
        backend_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 12px;")
        layout.addWidget(backend_label)
        
        self.backend_combo = QComboBox()
        self.backend_combo.setFixedHeight(36)
        self.backend_combo.addItems(["🖥️ Ollama (Local)", "☁️ Groq (Cloud)", "🤖 OpenAI (Cloud)", "🔮 Anthropic (Cloud)"])
        self.backend_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 13px;
            }}
        """)
        self.backend_combo.currentIndexChanged.connect(self._on_backend_change)
        layout.addWidget(self.backend_combo)

        # Ollama group
        self.ollama_group = QGroupBox("🦙 Ollama Settings")
        self.ollama_group.setStyleSheet(f"""
            QGroupBox {{
                color: {PALETTE['accent2']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 8px;
                font-weight: bold;
            }}
        """)
        og = QVBoxLayout(self.ollama_group)
        og.setSpacing(8)
        
        model_label = QLabel("Model:")
        model_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        og.addWidget(model_label)
        
        self.ollama_model = QComboBox()
        self.ollama_model.setFixedHeight(32)
        self.ollama_model.setEditable(True)
        self.ollama_model.addItems(["Loading..."])
        self.ollama_model.setStyleSheet(f"""
            QComboBox {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['accent']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        og.addWidget(self.ollama_model)

        def _fetch_models():
            try:
                import ollama
                models = ollama.list()
                model_names = [m.model for m in models.models] if hasattr(models, 'models') else []
                if model_names:
                    self.ollama_model.clear()
                    self.ollama_model.addItems(model_names)
                    self.ollama_status.setText("✅ Online")
                    self.ollama_status.setStyleSheet(f"color: {PALETTE['green']}; font-size: 12px; font-weight: bold;")
                else:
                    self.ollama_status.setText("⚠️ No models")
            except Exception as e:
                self.ollama_status.setText("❌ Offline")
                self.ollama_status.setStyleSheet(f"color: {PALETTE['red']}; font-size: 12px; font-weight: bold;")

        threading.Thread(target=_fetch_models, daemon=True).start()
        
        url_label = QLabel("Server URL:")
        url_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        og.addWidget(url_label)
        
        self.ollama_url = QLineEdit("http://localhost:11434")
        self.ollama_url.setFixedHeight(32)
        self.ollama_url.setStyleSheet(f"""
            QLineEdit {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        og.addWidget(self.ollama_url)

        self.ollama_status = QLabel("⏳ Connecting...")
        self.ollama_status.setStyleSheet(f"color: {PALETTE['warn']}; font-size: 12px; font-weight: bold; padding: 8px; background: {PALETTE['bg1']}; border-radius: 6px;")
        og.addWidget(self.ollama_status)
        layout.addWidget(self.ollama_group)

        # Groq group
        self.groq_group = QGroupBox("🔥 Groq Settings")
        self.groq_group.setStyleSheet(f"""
            QGroupBox {{
                color: {PALETTE['accent3']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 8px;
                font-weight: bold;
            }}
        """)
        gg = QVBoxLayout(self.groq_group)
        gg.setSpacing(8)
        
        gmodel_label = QLabel("Model:")
        gmodel_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        gg.addWidget(gmodel_label)
        
        self.groq_model = QComboBox()
        self.groq_model.setFixedHeight(32)
        self.groq_model.addItems([
            "llama-3.3-70b-versatile",
            "llama-3.1-8b-instant",
            "llama3-groq-70b-8192-tool-use-preview",
            "llama3-groq-8b-8192-tool-use-preview",
            "mixtral-8x7b-32768"
        ])
        self.groq_model.setStyleSheet(f"""
            QComboBox {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        gg.addWidget(self.groq_model)
        
        key_label = QLabel("API Key:")
        key_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        gg.addWidget(key_label)
        
        self.groq_key = QLineEdit()
        self.groq_key.setFixedHeight(32)
        self.groq_key.setPlaceholderText("gsk_...")
        self.groq_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.groq_key.setStyleSheet(f"""
            QLineEdit {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        self.groq_key.textChanged.connect(self._save_keys)
        gg.addWidget(self.groq_key)
        
        self.groq_status = QLabel("⏳ Waiting...")
        self.groq_status.setStyleSheet(f"color: {PALETTE['text3']}; font-size: 12px; font-weight: bold; padding: 8px; background: {PALETTE['bg1']}; border-radius: 6px;")
        gg.addWidget(self.groq_status)
        layout.addWidget(self.groq_group)

        # OpenAI group
        self.openai_group = QGroupBox("🤖 OpenAI Settings")
        self.openai_group.setStyleSheet(f"""
            QGroupBox {{
                color: {PALETTE['accent']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 8px;
                font-weight: bold;
            }}
        """)
        og2 = QVBoxLayout(self.openai_group)
        og2.setSpacing(8)

        oai_model_label = QLabel("Model:")
        oai_model_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        og2.addWidget(oai_model_label)

        self.openai_model = QComboBox()
        self.openai_model.setFixedHeight(32)
        self.openai_model.addItems(["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"])
        self.openai_model.setStyleSheet(f"""
            QComboBox {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        og2.addWidget(self.openai_model)

        oai_key_label = QLabel("API Key:")
        oai_key_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        og2.addWidget(oai_key_label)

        self.openai_key = QLineEdit()
        self.openai_key.setFixedHeight(32)
        self.openai_key.setPlaceholderText("sk-...")
        self.openai_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.openai_key.setStyleSheet(f"""
            QLineEdit {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        self.openai_key.textChanged.connect(self._save_keys)
        og2.addWidget(self.openai_key)

        self.openai_status = QLabel("⏳ Waiting...")
        self.openai_status.setStyleSheet(f"color: {PALETTE['text3']}; font-size: 12px; font-weight: bold; padding: 8px; background: {PALETTE['bg1']}; border-radius: 6px;")
        og2.addWidget(self.openai_status)
        layout.addWidget(self.openai_group)

        # Anthropic group
        self.anthropic_group = QGroupBox("🔮 Anthropic Settings")
        self.anthropic_group.setStyleSheet(f"""
            QGroupBox {{
                color: {PALETTE['magenta']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 8px;
                margin-top: 8px;
                padding-top: 8px;
                font-weight: bold;
            }}
        """)
        ag = QVBoxLayout(self.anthropic_group)
        ag.setSpacing(8)

        ant_model_label = QLabel("Model:")
        ant_model_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        ag.addWidget(ant_model_label)

        self.anthropic_model = QComboBox()
        self.anthropic_model.setFixedHeight(32)
        self.anthropic_model.addItems(["claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022", "claude-3-opus-20240229"])
        self.anthropic_model.setStyleSheet(f"""
            QComboBox {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        ag.addWidget(self.anthropic_model)

        ant_key_label = QLabel("API Key:")
        ant_key_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 11px;")
        ag.addWidget(ant_key_label)

        self.anthropic_key = QLineEdit()
        self.anthropic_key.setFixedHeight(32)
        self.anthropic_key.setPlaceholderText("sk-ant-...")
        self.anthropic_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.anthropic_key.setStyleSheet(f"""
            QLineEdit {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 8px;
            }}
        """)
        self.anthropic_key.textChanged.connect(self._save_keys)
        ag.addWidget(self.anthropic_key)

        self.anthropic_status = QLabel("⏳ Waiting...")
        self.anthropic_status.setStyleSheet(f"color: {PALETTE['text3']}; font-size: 12px; font-weight: bold; padding: 8px; background: {PALETTE['bg1']}; border-radius: 6px;")
        ag.addWidget(self.anthropic_status)
        layout.addWidget(self.anthropic_group)

        # Test button
        test_btn = QPushButton("🔗 Test Connection")
        test_btn.setFixedHeight(40)
        test_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        test_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {PALETTE['accent2']};
                color: {PALETTE['bg0']};
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: {PALETTE['cyan']};
            }}
        """)
        test_btn.clicked.connect(self._test_connection)
        layout.addWidget(test_btn)

        # Stats display
        stats_frame = QFrame()
        stats_frame.setStyleSheet(f"background: {PALETTE['bg1']}; border-radius: 8px; padding: 8px;")
        sf = QVBoxLayout(stats_frame)
        sf.setSpacing(4)
        
        stats_label = QLabel("📊 Statistics")
        stats_label.setStyleSheet(f"color: {PALETTE['accent2']}; font-size: 12px; font-weight: bold;")
        sf.addWidget(stats_label)
        
        self.stat_latency = QLabel("⏱️ Latency: --ms")
        self.stat_tps = QLabel("🚀 Speed: -- tok/s")
        self.stat_calls = QLabel("📝 Calls: 0")
        for lbl in [self.stat_latency, self.stat_tps, self.stat_calls]:
            lbl.setStyleSheet(f"color: {PALETTE['text']}; font-size: 11px; padding: 4px;")
            sf.addWidget(lbl)
        layout.addWidget(stats_frame)
        layout.addStretch()

        self._on_backend_change(0)

    def _on_backend_change(self, idx):
        self.ollama_group.setVisible(idx == 0)
        self.groq_group.setVisible(idx == 1)
        self.openai_group.setVisible(idx == 2)
        self.anthropic_group.setVisible(idx == 3)

    def _get_config(self) -> dict:
        backends = ["ollama", "groq", "openai", "anthropic"]
        idx = self.backend_combo.currentIndex()
        return {
            "backend": backends[idx],
            "ollama_model": self.ollama_model.currentText(),
            "ollama_url": self.ollama_url.text(),
            "groq_model": self.groq_model.currentText(),
            "groq_api_key": self.groq_key.text() or os.environ.get("GROQ_API_KEY", ""),
            "openai_model": self.openai_model.currentText(),
            "openai_api_key": self.openai_key.text() or os.environ.get("OPENAI_API_KEY", ""),
            "anthropic_model": self.anthropic_model.currentText(),
            "anthropic_api_key": self.anthropic_key.text() or os.environ.get("ANTHROPIC_API_KEY", ""),
        }

    def get_config(self) -> dict:
        return self._get_config()

    @staticmethod
    def _keys_path() -> str:
        d = os.environ.get("XDG_CONFIG_HOME", os.path.expanduser("~/.config"))
        p = os.path.join(d, "synthgrid")
        os.makedirs(p, exist_ok=True)
        return os.path.join(p, "keys.json")

    def _load_keys(self):
        try:
            with open(self._keys_path()) as f:
                keys = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return
        if "groq" in keys:
            self.groq_key.setText(keys["groq"])
        if "openai" in keys:
            self.openai_key.setText(keys["openai"])
        if "anthropic" in keys:
            self.anthropic_key.setText(keys["anthropic"])

    def _save_keys(self):
        keys = {
            "groq": self.groq_key.text(),
            "openai": self.openai_key.text(),
            "anthropic": self.anthropic_key.text(),
        }
        with open(self._keys_path(), "w") as f:
            json.dump(keys, f)

    def _test_connection(self):
        config = self._get_config()
        b = config["backend"]
        if b == "ollama":
            self._test_ollama(config)
        elif b == "groq":
            self._test_groq(config)
        elif b == "openai":
            self._test_openai(config)
        elif b == "anthropic":
            self._test_anthropic(config)

    def _test_ollama(self, config):
        def _run():
            try:
                import ollama
                t0 = time.time()
                models = ollama.list()
                elapsed = (time.time() - t0) * 1000
                model_names = [m.model for m in models.models] if hasattr(models, 'models') else []
                self.ollama_status.setText(f"● ONLINE  {elapsed:.0f}ms")
                self.ollama_status.setStyleSheet(f"color: {PALETTE['green']}; font-size: 10px;")
                self.stat_latency.setText(f"LATENCY: {elapsed:.0f}ms")
                # Populate model list
                if model_names:
                    current = self.ollama_model.currentText()
                    self.ollama_model.clear()
                    self.ollama_model.addItems(model_names)
                    idx = self.ollama_model.findText(current)
                    if idx >= 0:
                        self.ollama_model.setCurrentIndex(idx)
            except Exception as e:
                self.ollama_status.setText(f"● ERROR: {str(e)[:30]}")
                self.ollama_status.setStyleSheet(f"color: {PALETTE['red']}; font-size: 10px;")
        threading.Thread(target=_run, daemon=True).start()

    def _test_groq(self, config):
        def _run():
            try:
                from groq import Groq
                key = config["groq_api_key"]
                if not key:
                    self.groq_status.setText("● NO KEY SET")
                    return
                client = Groq(api_key=key)
                t0 = time.time()
                client.models.list()
                elapsed = (time.time() - t0) * 1000
                self.groq_status.setText(f"● ONLINE  {elapsed:.0f}ms")
                self.groq_status.setStyleSheet(f"color: {PALETTE['green']}; font-size: 10px;")
                self.stat_latency.setText(f"LATENCY: {elapsed:.0f}ms")
            except Exception as e:
                self.groq_status.setText(f"● ERROR: {str(e)[:30]}")
                self.groq_status.setStyleSheet(f"color: {PALETTE['red']}; font-size: 10px;")
        threading.Thread(target=_run, daemon=True).start()

    def _test_openai(self, config):
        def _run():
            try:
                from openai import OpenAI
                key = config["openai_api_key"]
                if not key:
                    self.openai_status.setText("● NO KEY SET")
                    return
                client = OpenAI(api_key=key)
                t0 = time.time()
                client.models.list()
                elapsed = (time.time() - t0) * 1000
                self.openai_status.setText(f"● ONLINE  {elapsed:.0f}ms")
                self.openai_status.setStyleSheet(f"color: {PALETTE['green']}; font-size: 10px;")
                self.stat_latency.setText(f"LATENCY: {elapsed:.0f}ms")
            except Exception as e:
                self.openai_status.setText(f"● ERROR: {str(e)[:30]}")
                self.openai_status.setStyleSheet(f"color: {PALETTE['red']}; font-size: 10px;")
        threading.Thread(target=_run, daemon=True).start()

    def _test_anthropic(self, config):
        def _run():
            try:
                from anthropic import Anthropic
                key = config["anthropic_api_key"]
                if not key:
                    self.anthropic_status.setText("● NO KEY SET")
                    return
                client = Anthropic(api_key=key)
                t0 = time.time()
                client.models.list()
                elapsed = (time.time() - t0) * 1000
                self.anthropic_status.setText(f"● ONLINE  {elapsed:.0f}ms")
                self.anthropic_status.setStyleSheet(f"color: {PALETTE['green']}; font-size: 10px;")
                self.stat_latency.setText(f"LATENCY: {elapsed:.0f}ms")
            except Exception as e:
                self.anthropic_status.setText(f"● ERROR: {str(e)[:30]}")
                self.anthropic_status.setStyleSheet(f"color: {PALETTE['red']}; font-size: 10px;")
        threading.Thread(target=_run, daemon=True).start()

    def update_stats(self, latency_ms: float = None, tps: float = None, calls: int = None):
        if latency_ms is not None:
            self.stat_latency.setText(f"LATENCY: {latency_ms:.0f}ms")
        if tps is not None:
            self.stat_tps.setText(f"TOK/S:   {tps:.1f}")
        if calls is not None:
            self.stat_calls.setText(f"CALLS:   {calls}")


class RAGPanel(QWidget):
    file_analyzed = pyqtSignal(dict)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.analyzer = RAGAnalyzer()
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(12, 8, 12, 8)

        header = QLabel("📁 Upload Excel File")
        header.setStyleSheet(f"color: {PALETTE['accent2']}; font-size: 14px; font-weight: bold;")
        layout.addWidget(header)

        self.drop_zone = QFrame()
        self.drop_zone.setCursor(Qt.CursorShape.PointingHandCursor)
        self.drop_zone.setStyleSheet(f"""
            QFrame {{
                background-color: {PALETTE['bg2']};
                border: 2px dashed {PALETTE['border2']};
                border-radius: 12px;
            }}
            QFrame:hover {{
                border-color: {PALETTE['accent']};
                background-color: {PALETTE['bg3']};
            }}
        """)
        self.drop_zone.setFixedHeight(100)
        dz_layout = QVBoxLayout(self.drop_zone)
        dz_layout.setSpacing(8)
        
        self.drop_icon = QLabel("📂")
        self.drop_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drop_icon.setStyleSheet(f"font-size: 32px;")
        dz_layout.addWidget(self.drop_icon)
        
        self.drop_label = QLabel("Click to select Excel file")
        self.drop_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drop_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 12px;")
        dz_layout.addWidget(self.drop_label)
        
        self.drop_zone.mousePressEvent = self._on_drop_click
        layout.addWidget(self.drop_zone)

        self.analysis_label = QLabel("No file loaded")
        self.analysis_label.setStyleSheet(f"color: {PALETTE['text3']}; font-size: 12px; padding: 8px; background: {PALETTE['bg1']}; border-radius: 6px;")
        self.analysis_label.setWordWrap(True)
        layout.addWidget(self.analysis_label)

        self.sheet_info = QLabel("")
        self.sheet_info.setStyleSheet(f"color: {PALETTE['accent2']}; font-size: 11px;")
        self.sheet_info.setWordWrap(True)
        layout.addWidget(self.sheet_info)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        
        self.analyze_btn = QPushButton("🔍 Analyze")
        self.analyze_btn.setFixedHeight(36)
        self.analyze_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.analyze_btn.setEnabled(False)
        self.analyze_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {PALETTE['accent']};
                color: {PALETTE['bg0']};
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {PALETTE['green']};
            }}
            QPushButton:disabled {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text3']};
            }}
        """)
        btn_row.addWidget(self.analyze_btn)

        self.clear_btn = QPushButton("🗑️ Clear")
        self.clear_btn.setFixedHeight(36)
        self.clear_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.clear_btn.setEnabled(False)
        self.clear_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text2']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 8px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                border-color: {PALETTE['red']};
                color: {PALETTE['red']};
            }}
        """)
        btn_row.addWidget(self.clear_btn)
        
        layout.addLayout(btn_row)

        layout.addStretch()

        self.analyze_btn.clicked.connect(self._run_analysis)
        self.clear_btn.clicked.connect(self._clear_file)

    def _on_drop_click(self, event):
        self._open_file_dialog()

    def _open_file_dialog(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm *.xls);;All Files (*)"
        )
        if path:
            self._load_file(path)

    def _load_file(self, filepath: str):
        result = self.analyzer.analyze_file(filepath)
        if result["status"] == "ok":
            fname = os.path.basename(filepath)
            self.analysis_label.setText(f"◈ LOADED: {fname}")
            self.analysis_label.setStyleSheet(f"color: {PALETTE['accent']}; font-size: 10px;")
            
            sheets_info = []
            for sheet, info in result["schema"].items():
                sheets_info.append(f"{sheet}: {info['rows']} rows × {len(info['headers'])} cols")
            self.sheet_info.setText(" | ".join(sheets_info))
            
            self.analyze_btn.setEnabled(True)
            self.clear_btn.setEnabled(True)
            
            self.drop_label.setText("File loaded - click Analyze or Ask AI what to do")
            self.drop_icon.setStyleSheet(f"color: {PALETTE['accent']}; font-size: 24px;")
            
            self.file_analyzed.emit(result)
        else:
            self.analysis_label.setText(f"✗ {result['message']}")
            self.analysis_label.setStyleSheet(f"color: {PALETTE['red']}; font-size: 10px;")

    def _run_analysis(self):
        if self.analyzer.current_file:
            result = self.analyzer.analyze_file(self.analyzer.current_file)
            self.file_analyzed.emit(result)

    def _clear_file(self):
        self.analyzer.clear()
        self.analysis_label.setText("◈ NO FILE LOADED")
        self.analysis_label.setStyleSheet(f"color: {PALETTE['text3']}; font-size: 10px;")
        self.sheet_info.setText("")
        self.analyze_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        self.drop_label.setText("Drop Excel file here or click to browse")
        self.drop_icon.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 24px;")

    def get_context(self) -> str:
        return self.analyzer.get_context_prompt()

    def has_file(self) -> bool:
        return self.analyzer.current_file is not None


class ContinueWorker(QThread):
    log_signal = pyqtSignal(str, str)
    grid_signal = pyqtSignal()
    done_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

    def __init__(self, prompt: str, engine: ExcelEngine, model_config: dict, previous_messages: list):
        super().__init__()
        self.prompt = prompt
        self.engine = engine
        self.model_config = model_config
        self._messages = previous_messages.copy()
        self._messages.append({"role": "user", "content": prompt})

    def log(self, msg: str, level: str = "info"):
        self.log_signal.emit(msg, level)

    def execute_tool(self, name: str, args: dict) -> str:
        method = getattr(self.engine, name, None)
        if not method:
            return json.dumps({"status": "error", "message": f"Unknown tool: {name}"})
        try:
            result = method(**args)
            self.grid_signal.emit()
            return json.dumps(result)
        except Exception as e:
            return json.dumps({"status": "error", "message": str(e), "traceback": traceback.format_exc()})

    def run_with_groq(self, messages):
        try:
            from groq import Groq
            api_key = self.model_config.get("groq_api_key", os.environ.get("GROQ_API_KEY", ""))
            if not api_key:
                raise ValueError("GROQ_API_KEY not set")
            client = Groq(api_key=api_key)
            model = self.model_config.get("groq_model", "llama-3.3-70b-versatile")
            self.log(f"◈ GROQ CONTINUE REQUEST → model={model}", "api")
            t0 = time.time()
            resp = client.chat.completions.create(
                model=model,
                messages=messages,
                tools=TOOL_SCHEMAS,
                tool_choice="auto",
                max_tokens=4096,
            )
            elapsed = time.time() - t0
            usage = resp.usage
            tps = (usage.completion_tokens / elapsed) if elapsed > 0 else 0
            self.log(f"◈ GROQ RESPONSE ← {usage.completion_tokens} tokens | {tps:.1f} tok/s | {elapsed:.2f}s", "api")
            return resp.choices[0].message
        except Exception as e:
            raise RuntimeError(f"Groq error: {e}")

    def run_with_ollama(self, messages):
        try:
            import ollama
            model = self.model_config.get("ollama_model", "llama3.1")
            self.log(f"◈ OLLAMA CONTINUE REQUEST → model={model}", "api")
            t0 = time.time()
            tools = [{"type": "function", "function": t["function"]} for t in TOOL_SCHEMAS]
            resp = ollama.chat(model=model, messages=messages, tools=tools, stream=False)
            elapsed = time.time() - t0
            self.log(f"◈ OLLAMA RESPONSE ← {elapsed:.2f}s", "api")
            return resp.message
        except Exception as e:
            raise RuntimeError(f"Ollama error: {e}")

    def run_with_openai(self, messages):
        try:
            from openai import OpenAI
            api_key = self.model_config.get("openai_api_key", os.environ.get("OPENAI_API_KEY", ""))
            if not api_key:
                raise ValueError("OPENAI_API_KEY not set")
            client = OpenAI(api_key=api_key)
            model = self.model_config.get("openai_model", "gpt-4o")
            self.log(f"◈ OPENAI CONTINUE REQUEST → model={model}", "api")
            t0 = time.time()
            resp = client.chat.completions.create(
                model=model, messages=messages, tools=TOOL_SCHEMAS,
                tool_choice="auto", max_tokens=4096,
            )
            elapsed = time.time() - t0
            u = resp.usage
            tps = (u.completion_tokens / elapsed) if elapsed > 0 else 0
            self.log(f"◈ OPENAI RESPONSE ← {u.completion_tokens} tokens | {tps:.1f} tok/s | {elapsed:.2f}s", "api")
            return resp.choices[0].message
        except Exception as e:
            raise RuntimeError(f"OpenAI error: {e}")

    def run_with_anthropic(self, messages):
        try:
            from anthropic import Anthropic
            api_key = self.model_config.get("anthropic_api_key", os.environ.get("ANTHROPIC_API_KEY", ""))
            if not api_key:
                raise ValueError("ANTHROPIC_API_KEY not set")
            client = Anthropic(api_key=api_key)
            model = self.model_config.get("anthropic_model", "claude-3-5-sonnet-20241022")
            self.log(f"◈ ANTHROPIC CONTINUE REQUEST → model={model}", "api")
            t0 = time.time()

            ant_tools = [{"name": t["function"]["name"], "description": t["function"].get("description", ""), "input_schema": t["function"]["parameters"]} for t in TOOL_SCHEMAS]
            system = next((m["content"] for m in messages if m["role"] == "system"), "")

            ant_messages = []
            for m in messages:
                r = m["role"]
                if r == "system":
                    continue
                if r == "tool":
                    ant_messages.append({"role": "user", "content": [{"type": "tool_result", "tool_use_id": m["tool_call_id"], "content": m["content"]}]})
                elif r == "assistant" and "tool_calls" in m:
                    c = []
                    if m.get("content"):
                        c.append({"type": "text", "text": m["content"]})
                    for tc in m["tool_calls"]:
                        c.append({"type": "tool_use", "id": tc["id"], "name": tc["function"]["name"], "input": tc["function"]["arguments"]})
                    ant_messages.append({"role": "assistant", "content": c})
                else:
                    ant_messages.append({"role": r, "content": m.get("content", "")})

            resp = client.messages.create(model=model, system=system, messages=ant_messages, tools=ant_tools or None, max_tokens=4096)
            elapsed = time.time() - t0

            content = ""
            tool_calls = []
            for block in resp.content:
                if block.type == "text":
                    content = block.text
                elif block.type == "tool_use":
                    tool_calls.append(type('ToolCall', (), {'id': block.id, 'function': type('Fn', (), {'name': block.name, 'arguments': json.dumps(block.input)})})())

            msg = type('Msg', (), {'content': content, 'tool_calls': tool_calls or None})()
            self.log(f"◈ ANTHROPIC RESPONSE ← {elapsed:.2f}s", "api")
            return msg
        except Exception as e:
            raise RuntimeError(f"Anthropic error: {e}")

    def run(self):
        try:
            self.log(f"▶ CONTINUING CONVERSATION: {self.prompt[:60]}...", "start")
            backend = self.model_config.get("backend", "ollama")
            max_iterations = 10

            for iteration in range(max_iterations):
                self.log(f"── CONTINUE ITERATION {iteration + 1} ──────────────────", "divider")
                
                if backend == "groq":
                    message = self.run_with_groq(self._messages)
                elif backend == "openai":
                    message = self.run_with_openai(self._messages)
                elif backend == "anthropic":
                    message = self.run_with_anthropic(self._messages)
                else:
                    message = self.run_with_ollama(self._messages)

                tool_calls = getattr(message, "tool_calls", None)
                
                if not tool_calls:
                    content = getattr(message, "content", "") or ""
                    self.log(f"✓ AGENT FINAL: {content}", "success")
                    self.done_signal.emit(content)
                    return

                msg_dict = {"role": "assistant", "content": getattr(message, "content", "") or ""}
                if tool_calls:
                    msg_dict["tool_calls"] = [
                        {
                            "id": tc.id if hasattr(tc, 'id') else f"call_{i}",
                            "type": "function",
                            "function": {
                                "name": tc.function.name,
                                "arguments": json.loads(tc.function.arguments) if isinstance(tc.function.arguments, str) else tc.function.arguments
                            }
                        }
                        for i, tc in enumerate(tool_calls)
                    ]
                self._messages.append(msg_dict)

                for tc in tool_calls:
                    fn_name = tc.function.name
                    fn_args_raw = tc.function.arguments
                    tc_id = tc.id if hasattr(tc, 'id') else f"call_0"
                    
                    if isinstance(fn_args_raw, str):
                        try:
                            fn_args = json.loads(fn_args_raw)
                        except:
                            fn_args = {}
                    else:
                        fn_args = fn_args_raw or {}

                    self.log(f"⚡ TOOL CALL → {fn_name}({json.dumps(fn_args)[:120]})", "tool")
                    result_str = self.execute_tool(fn_name, fn_args)
                    result_obj = json.loads(result_str)
                    
                    status = result_obj.get("status", "?")
                    if status == "ok":
                        msg_part = result_obj.get("message", result_str[:100])
                        self.log(f"  ✓ RESULT: {msg_part}", "ok")
                    else:
                        self.log(f"  ✗ ERROR: {result_obj.get('message', result_str[:100])}", "error")

                    self._messages.append({
                        "role": "tool",
                        "tool_call_id": tc_id,
                        "content": result_str
                    })

            self.log("⚠ Max iterations reached", "warn")
            self.done_signal.emit("Task completed (max iterations reached)")
        except Exception as e:
            self.error_signal.emit(str(e))
            self.log(f"✗ FATAL: {e}", "error")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN WINDOW
# ─────────────────────────────────────────────────────────────────────────────

ASCII_BANNER = """
╔══════════════════════════════════════════════════════════════════════════════════╗
║  ▓▓▓  S Y N T H G R I D  ▓▓▓   AI-POWERED EXCEL AUTOMATION NODE  v1.0.0       ║
║  ◈ BACKEND: OLLAMA/GROQ  ◈ ENGINE: OPENPYXL  ◈ MODE: CYBERSIGILISM INDUSTRIAL ║
╚══════════════════════════════════════════════════════════════════════════════════╝"""


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.engine = ExcelEngine()
        self.agent_worker: Optional[AgentWorker] = None
        self.call_count = 0
        self.conversation_history = []
        self.last_messages = None
        self.is_paused = False
        self.memory = MemoryManager()
        self._build_ui()
        self._connect_signals()
        self._post_init()

    def _build_ui(self):
        self.setWindowTitle("SYNTHGRID — AI Excel Automation Node")
        self.setMinimumSize(1400, 900)
        self.resize(1600, 1000)
        self.setStyleSheet(GLOBAL_STYLE)

        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(0)
        root.setContentsMargins(4, 4, 4, 0)

        # ── TOP BANNER ─────────────────────────────────────────────────────
        banner = QLabel(ASCII_BANNER)
        banner.setStyleSheet(f"""
            color: {PALETTE['accent']};
            background-color: {PALETTE['bg1']};
            font-family: 'JetBrains Mono', monospace;
            font-size: 10px;
            padding: 4px 8px;
            border-bottom: 1px solid {PALETTE['border2']};
        """)
        banner.setAlignment(Qt.AlignmentFlag.AlignLeft)
        root.addWidget(banner)

        # ── TOOLBAR ───────────────────────────────────────────────────────
        toolbar = self._build_toolbar()
        root.addWidget(toolbar)

        # ── MAIN SPLITTER ─────────────────────────────────────────────────
        main_split = QSplitter(Qt.Orientation.Horizontal)
        main_split.setHandleWidth(2)

        # LEFT: Node Selector + RAG Panel
        left_panel = self._build_left_panel()
        left_panel.setMinimumWidth(240)
        left_panel.setMaximumWidth(320)
        main_split.addWidget(left_panel)
        
        self.rag_panel.file_analyzed.connect(self._on_rag_file_analyzed)

        # CENTER: Main workspace
        center = self._build_center()
        main_split.addWidget(center)

        main_split.setSizes([260, 1340])
        root.addWidget(main_split, 1)

        # ── STATUS BAR ────────────────────────────────────────────────────
        self._build_statusbar()

    def _build_toolbar(self) -> QWidget:
        bar = QFrame()
        bar.setStyleSheet(f"""
            QFrame {{
                background-color: {PALETTE['bg1']};
                border-bottom: 2px solid {PALETTE['accent']}44;
            }}
        """)
        bar.setFixedHeight(56)
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(12, 6, 12, 6)
        layout.setSpacing(8)

        def btn(text, tooltip="", obj_name="", icon=""):
            b = QPushButton(f"{icon} {text}" if icon else text)
            b.setToolTip(tooltip)
            b.setFixedHeight(40)
            b.setMinimumWidth(80)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setStyleSheet(f"""
                QPushButton {{
                    background-color: {PALETTE['bg2']};
                    color: {PALETTE['accent']};
                    border: 1px solid {PALETTE['accent']}66;
                    border-radius: 6px;
                    padding: 4px 12px;
                    font-weight: bold;
                    font-size: 12px;
                }}
                QPushButton:hover {{
                    background-color: {PALETTE['accent']};
                    color: {PALETTE['bg0']};
                    border-color: {PALETTE['accent']};
                }}
                QPushButton:pressed {{
                    background-color: {PALETTE['accent']}CC;
                }}
            """)
            if obj_name:
                b.setObjectName(obj_name)
            return b

        def icon_btn(text, tooltip="", icon="📄"):
            return btn(text, tooltip, icon=icon)

        self.btn_new = icon_btn("New", "Create new workbook", "📄")
        self.btn_open = icon_btn("Open", "Open existing .xlsx", "📂")
        self.btn_save = icon_btn("Save", "Save workbook", "💾")
        self.btn_save_as = icon_btn("Save As", "Save as new file", "📥")
        
        sep = QLabel("│")
        sep.setStyleSheet(f"color: {PALETTE['border2']}; font-size: 20px;")

        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(150)
        self.sheet_combo.setFixedHeight(36)
        self.sheet_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['accent2']};
                border: 1px solid {PALETTE['border2']};
                border-radius: 6px;
                padding: 4px 12px;
                font-size: 12px;
            }}
            QComboBox:hover {{
                border-color: {PALETTE['accent']};
            }}
        """)
        self.sheet_combo.setToolTip("Active sheet")
        
        sheet_label = QLabel("📑 Sheet:")
        sheet_label.setStyleSheet(f"color: {PALETTE['text']}; font-size: 12px; font-weight: bold;")
        
        self.btn_add_sheet = btn("+", "Add new sheet", icon="➕")
        self.btn_add_sheet.setFixedWidth(40)
        self.btn_del_sheet = btn("−", "Delete active sheet", "danger")
        self.btn_del_sheet.setFixedWidth(40)
        self.btn_del_sheet.setStyleSheet(f"""
            QPushButton {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['red']};
                border: 1px solid {PALETTE['red']}66;
                border-radius: 6px;
                padding: 4px 8px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: {PALETTE['red']};
                color: {PALETTE['bg0']};
            }}
        """)

        sep2 = QLabel("│")
        sep2.setStyleSheet(f"color: {PALETTE['border2']}; font-size: 20px;")

        self.progress = QProgressBar()
        self.progress.setFixedWidth(150)
        self.progress.setFixedHeight(8)
        self.progress.setRange(0, 0)
        self.progress.setVisible(False)
        self.progress.setStyleSheet(f"""
            QProgressBar {{
                background-color: {PALETTE['bg2']};
                border: none;
                border-radius: 4px;
            }}
            QProgressBar::chunk {{
                background-color: {PALETTE['accent']};
                border-radius: 4px;
            }}
        """)

        self.agent_label = QLabel("◉ READY")
        self.agent_label.setStyleSheet(f"color: {PALETTE['green']}; font-size: 12px; font-weight: bold; font-family: monospace; padding: 8px 12px; background: {PALETTE['bg2']}; border-radius: 6px;")

        layout.addWidget(self.btn_new)
        layout.addWidget(self.btn_open)
        layout.addWidget(self.btn_save)
        layout.addWidget(self.btn_save_as)
        layout.addWidget(sep)
        layout.addWidget(sheet_label)
        layout.addWidget(self.sheet_combo)
        layout.addWidget(self.btn_add_sheet)
        layout.addWidget(self.btn_del_sheet)
        layout.addWidget(sep2)
        layout.addStretch()
        layout.addWidget(self.progress)
        layout.addWidget(self.agent_label)
        return bar

    def _build_left_panel(self) -> QWidget:
        panel = QFrame()
        panel.setObjectName("panel")
        layout = QVBoxLayout(panel)
        layout.setSpacing(0)
        layout.setContentsMargins(0, 0, 0, 0)
        
        layout.addWidget(SectionHeader("NODE SELECTOR"))
        self.node_selector = NodeSelector()
        layout.addWidget(self.node_selector, 1)
        
        layout.addWidget(SectionHeader("RAG ◈ FILE UPLOAD"))
        self.rag_panel = RAGPanel()
        layout.addWidget(self.rag_panel, 1)
        
        return panel

    def _build_center(self) -> QWidget:
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(2)
        layout.setContentsMargins(0, 0, 0, 0)

        # ── TOP HALF: Log + Grid split ─────────────────────────────────
        top_split = QSplitter(Qt.Orientation.Horizontal)
        top_split.setHandleWidth(2)

        # Neural Uplink (log)
        log_panel = QFrame()
        log_panel.setObjectName("panel")
        lp_layout = QVBoxLayout(log_panel)
        lp_layout.setSpacing(0)
        lp_layout.setContentsMargins(0, 0, 0, 0)
        lp_layout.addWidget(SectionHeader("NEURAL UPLINK ◈ AGENT VISIBILITY"))
        self.log_console = LogConsole()
        lp_layout.addWidget(self.log_console, 1)

        log_ctrl = QHBoxLayout()
        log_ctrl.setContentsMargins(4, 2, 4, 2)
        self.btn_clear_log = QPushButton("CLEAR LOG")
        self.btn_clear_log.setFixedHeight(20)
        self.btn_copy_log = QPushButton("COPY LOG")
        self.btn_copy_log.setFixedHeight(20)
        log_ctrl.addWidget(self.btn_clear_log)
        log_ctrl.addWidget(self.btn_copy_log)
        log_ctrl.addStretch()
        log_panel.layout().addLayout(log_ctrl)
        top_split.addWidget(log_panel)

        # Grid Render
        grid_panel = QFrame()
        grid_panel.setObjectName("panel")
        gp_layout = QVBoxLayout(grid_panel)
        gp_layout.setSpacing(0)
        gp_layout.setContentsMargins(0, 0, 0, 0)
        gp_layout.addWidget(SectionHeader("GRID RENDER ◈ DATA PREVIEW"))
        self.grid_view = GridView()
        gp_layout.addWidget(self.grid_view, 1)

        grid_info = QHBoxLayout()
        grid_info.setContentsMargins(4, 2, 4, 2)
        self.grid_info_label = QLabel("◈ NO DATA")
        self.grid_info_label.setStyleSheet(f"color: {PALETTE['text3']}; font-size: 10px;")
        self.btn_refresh_grid = QPushButton("◈ REFRESH")
        self.btn_refresh_grid.setFixedHeight(20)
        grid_info.addWidget(self.grid_info_label)
        grid_info.addStretch()
        grid_info.addWidget(self.btn_refresh_grid)
        gp_layout.addLayout(grid_info)
        top_split.addWidget(grid_panel)

        top_split.setSizes([700, 630])
        layout.addWidget(top_split, 1)

        # ── COMMAND MATRIX (Input area) ─────────────────────────────────
        cmd_panel = QFrame()
        cmd_panel.setObjectName("panel")
        cmd_panel.setFixedHeight(140)
        cmd_panel.setStyleSheet(f"QFrame {{ border-top: 2px solid {PALETTE['accent']}44; }}")
        cp_layout = QVBoxLayout(cmd_panel)
        cp_layout.setSpacing(8)
        cp_layout.setContentsMargins(12, 8, 12, 8)
        
        header = QLabel("💬 Ask AI to work with your Excel file")
        header.setStyleSheet(f"color: {PALETTE['accent']}; font-size: 14px; font-weight: bold; padding: 4px 0;")
        cp_layout.addWidget(header)
        
        input_row = QHBoxLayout()
        input_row.setContentsMargins(0, 0, 0, 0)
        input_row.setSpacing(10)

        self.prompt_icon = QLabel("💡")
        self.prompt_icon.setStyleSheet(f"font-size: 24px;")
        
        self.cmd_input = QLineEdit()
        self.cmd_input.setPlaceholderText(
            "e.g., 'Create a sales report with charts and formulas' or 'Analyze this data and add a summary'"
        )
        self.cmd_input.setFixedHeight(48)
        self.cmd_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text']};
                border: 2px solid {PALETTE['border2']};
                border-radius: 8px;
                padding: 10px 16px;
                font-size: 14px;
            }}
            QLineEdit:focus {{
                border-color: {PALETTE['accent']};
                background-color: {PALETTE['bg3']};
            }}
            QLineEdit::placeholder {{
                color: {PALETTE['text3']};
            }}
        """)
        
        self.btn_execute = QPushButton("▶ RUN")
        self.btn_execute.setFixedSize(100, 48)
        self.btn_execute.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_execute.setStyleSheet(f"""
            QPushButton {{
                background-color: {PALETTE['accent']};
                color: {PALETTE['bg0']};
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 14px;
                letter-spacing: 1px;
            }}
            QPushButton:hover {{
                background-color: {PALETTE['green']};
            }}
            QPushButton:pressed {{
                background-color: {PALETTE['accent']}CC;
            }}
            QPushButton:disabled {{
                background-color: {PALETTE['bg2']};
                color: {PALETTE['text3']};
            }}
        """)
        
        self.btn_stop = QPushButton("⏹ STOP")
        self.btn_stop.setFixedSize(90, 48)
        self.btn_stop.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_stop.setObjectName("danger")
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet(f"""
            QPushButton {{
                background-color: {PALETTE['red']};
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background-color: {PALETTE['accent3']};
            }}
        """)

        input_row.addWidget(self.prompt_icon)
        input_row.addWidget(self.cmd_input, 1)
        input_row.addWidget(self.btn_execute)
        input_row.addWidget(self.btn_stop)
        cp_layout.addLayout(input_row)

        # Quick commands
        quick_row = QHBoxLayout()
        quick_row.setContentsMargins(0, 8, 0, 0)
        quick_row.setSpacing(8)
        
        quick_label = QLabel("⚡ Quick:")
        quick_label.setStyleSheet(f"color: {PALETTE['text2']}; font-size: 12px; font-weight: bold;")
        quick_row.addWidget(quick_label)
        
        quick_cmds = [
            ("📊 Sales Report", "Create a Q3 sales projection with product data, SUM formulas, and a bar chart"),
            ("💰 Budget", "Create a monthly budget spreadsheet with income, expenses, and balance formulas"),
            ("📋 Inventory", "Create a sample inventory table with 10 items, quantities, prices, and calculated totals"),
            ("🎨 Format", "Make row 1 bold, white text, dark background, and freeze the top row"),
            ("📈 Chart", "Add a line chart to the current data"),
        ]
        for label, cmd in quick_cmds:
            b = QPushButton(label)
            b.setFixedHeight(32)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
            b.setStyleSheet(f"""
                QPushButton {{
                    font-size: 12px;
                    padding: 4px 14px;
                    color: {PALETTE['text']};
                    border: 1px solid {PALETTE['border2']};
                    border-radius: 16px;
                    background-color: {PALETTE['bg2']};
                }}
                QPushButton:hover {{
                    color: {PALETTE['accent']};
                    border-color: {PALETTE['accent']};
                    background-color: {PALETTE['accent']}22;
                }}
            """)
            b.clicked.connect(lambda _, c=cmd: self.cmd_input.setText(c))
            quick_row.addWidget(b)
        quick_row.addStretch()
        cp_layout.addLayout(quick_row)
        layout.addWidget(cmd_panel)

        return widget

    def _build_statusbar(self):
        sb = self.statusBar()
        self.status_file = QLabel("◈ UNTITLED")
        self.status_sheet = QLabel("SHEET: Sheet1")
        self.status_cells = QLabel("CELLS: 0")
        self.status_backend = QLabel("BACKEND: OLLAMA")
        self.status_time = QLabel("")
        
        for lbl in [self.status_file, self.status_sheet, self.status_cells, self.status_backend]:
            lbl.setStyleSheet(f"color: {PALETTE['text2']}; padding: 0 12px; border-right: 1px solid {PALETTE['border']};")
            sb.addWidget(lbl)
        sb.addPermanentWidget(self.status_time)
        
        # Clock timer
        self._clock_timer = QTimer()
        self._clock_timer.timeout.connect(self._update_clock)
        self._clock_timer.start(1000)
        self._update_clock()

    def _connect_signals(self):
        self.btn_new.clicked.connect(self._new_workbook)
        self.btn_open.clicked.connect(self._open_workbook)
        self.btn_save.clicked.connect(self._save_workbook)
        self.btn_save_as.clicked.connect(self._save_workbook_as)
        self.btn_add_sheet.clicked.connect(self._add_sheet)
        self.btn_del_sheet.clicked.connect(self._del_sheet)
        self.sheet_combo.currentTextChanged.connect(self._on_sheet_change)
        self.btn_execute.clicked.connect(self._execute_command)
        self.btn_stop.clicked.connect(self._stop_agent)
        self.cmd_input.returnPressed.connect(self._execute_command)
        self.btn_clear_log.clicked.connect(self.log_console.clear)
        self.btn_copy_log.clicked.connect(lambda: QApplication.clipboard().setText(self.log_console.toPlainText()))
        self.btn_refresh_grid.clicked.connect(self._refresh_grid)
        QShortcut(QKeySequence("Ctrl+Return"), self).activated.connect(self._execute_command)
        QShortcut(QKeySequence("Ctrl+S"), self).activated.connect(self._save_workbook)
        QShortcut(QKeySequence("Ctrl+O"), self).activated.connect(self._open_workbook)
        QShortcut(QKeySequence("Ctrl+N"), self).activated.connect(self._new_workbook)

    def _post_init(self):
        self.log_console.append_log(ASCII_BANNER, "success")
        self.log_console.append_log("SYNTHGRID INITIALIZED — EXCEL AUTOMATION NODE READY", "success")
        self.log_console.append_log("Configure AI backend in NODE SELECTOR panel (left)", "info")
        self.log_console.append_log("Type a command in COMMAND MATRIX and press EXECUTE or Enter", "info")
        self.log_console.append_log("Use RAG panel to upload Excel files and ask 'what to do'", "info")
        self.log_console.append_log("─" * 80, "divider")
        self._update_sheet_combo()
        self._refresh_grid()

    def _on_rag_file_analyzed(self, result: dict):
        if result["status"] == "ok":
            filepath = result["file"]
            open_result = self.engine.open_workbook(filepath)
            if open_result["status"] == "ok":
                self.log_console.append_log(f"◈ RAG: File loaded into engine", "ok")
                self._update_sheet_combo()
                self._refresh_grid()
                self._update_status()
            else:
                self.log_console.append_log(f"◈ RAG: {open_result['message']}", "error")

    # ── Workbook Actions ───────────────────────────────────────────────────

    def _new_workbook(self):
        self.engine._new_workbook()
        self.log_console.append_log("◈ NEW WORKBOOK CREATED", "ok")
        self._update_sheet_combo()
        self._refresh_grid()
        self._update_status()

    def _open_workbook(self):
        path, _ = QFileDialog.getOpenFileName(self, "Open Workbook", "", "Excel Files (*.xlsx *.xlsm);;All Files (*)")
        if path:
            result = self.engine.open_workbook(path)
            self.log_console.append_log(f"◈ {result['message']}", "ok" if result["status"] == "ok" else "error")
            self._update_sheet_combo()
            self._refresh_grid()
            self._update_status()

    def _save_workbook(self):
        if not self.engine.filepath:
            self._save_workbook_as()
            return
        result = self.engine.save_workbook()
        self.log_console.append_log(f"◈ {result['message']}", "ok" if result["status"] == "ok" else "error")
        self._update_status()

    def _save_workbook_as(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Workbook As", "workbook.xlsx", "Excel Files (*.xlsx)")
        if path:
            if not path.endswith(".xlsx"):
                path += ".xlsx"
            result = self.engine.save_workbook(path)
            self.log_console.append_log(f"◈ {result['message']}", "ok" if result["status"] == "ok" else "error")
            self._update_status()

    def _add_sheet(self):
        name = f"Sheet{len(self.engine.get_sheet_names())+1}"
        result = self.engine.create_sheet(name)
        self.log_console.append_log(f"◈ {result['message']}", "ok")
        self._update_sheet_combo()

    def _del_sheet(self):
        name = self.sheet_combo.currentText()
        if name:
            result = self.engine.delete_sheet(name)
            self.log_console.append_log(f"◈ {result['message']}", "ok" if result["status"] == "ok" else "error")
            self._update_sheet_combo()
            self._refresh_grid()

    def _on_sheet_change(self, name: str):
        if name and name in self.engine.get_sheet_names():
            self.engine.set_active_sheet(name)
            self._refresh_grid()
            self.status_sheet.setText(f"SHEET: {name}")

    def _update_sheet_combo(self):
        sheets = self.engine.get_sheet_names()
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(sheets)
        active = self.engine.active_sheet_name
        if active in sheets:
            self.sheet_combo.setCurrentText(active)
        self.sheet_combo.blockSignals(False)

    def _refresh_grid(self):
        data = self.engine.get_sheet_data()
        self.grid_view.load_data(data)
        rows = len(data)
        cols = max(len(r) for r in data) if data else 0
        self.grid_info_label.setText(f"◈ {rows}×{cols}  |  {self.engine.active_sheet_name or 'No sheet'}")

    def _update_status(self):
        fp = self.engine.filepath or "UNTITLED"
        self.status_file.setText(f"◈ {os.path.basename(fp)}")
        self.status_sheet.setText(f"SHEET: {self.engine.active_sheet_name or '-'}")
        config = self.node_selector.get_config()
        self.status_backend.setText(f"BACKEND: {config['backend'].upper()}")

    def _update_clock(self):
        self.status_time.setText(datetime.now().strftime("◈ %Y-%m-%d  %H:%M:%S"))

    # ── Agent Execution ────────────────────────────────────────────────────

    def _execute_command(self):
        prompt = self.cmd_input.text().strip()
        if not prompt:
            return
        if self.agent_worker and self.agent_worker.isRunning():
            self.log_console.append_log("⚠ AGENT BUSY — STOP FIRST", "warn")
            return

        config = self.node_selector.get_config()
        
        is_continue = prompt.lower() in ["continue", "continue task", "resume", "go on", "keep going", "continue please"]
        
        if is_continue and self.last_messages:
            self.log_console.append_log("◈ RESUMING PREVIOUS CONVERSATION...", "info")
            self._continue_agent(prompt, config)
            return

        if is_continue and not self.last_messages:
            self.log_console.append_log("⚠ No previous conversation to continue. Start a new task.", "warn")
            return

        self.call_count += 1
        rag_context = self.rag_panel.get_context() if hasattr(self, 'rag_panel') else ""
        
        self.log_console.append_log("─" * 80, "divider")
        self.log_console.append_log(f"▶ COMMAND #{self.call_count}: {prompt}", "start")
        self.log_console.append_log(f"  BACKEND: {config['backend'].upper()} | MODEL: {config.get(config['backend']+'_model', '?')}", "info")
        if rag_context:
            self.log_console.append_log("  ◈ RAG CONTEXT: ACTIVE", "info")
        if self.memory:
            self.log_console.append_log("  ◈ MEMORY: ACTIVE", "info")

        self.agent_worker = AgentWorker(prompt, self.engine, config, rag_context, self.memory)
        self.agent_worker.log_signal.connect(self.log_console.append_log)
        self.agent_worker.grid_signal.connect(self._on_grid_update)
        self.agent_worker.done_signal.connect(self._on_agent_done)
        self.agent_worker.error_signal.connect(self._on_agent_error)
        self.agent_worker.start()

        self.btn_execute.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.progress.setVisible(True)
        self.agent_label.setText("◉ AGENT RUNNING")
        self.agent_label.setStyleSheet(f"color: {PALETTE['yellow']}; font-size: 10px; font-family: monospace;")
        self.cmd_input.clear()
        self._update_status()
        self.is_paused = False

    def _continue_agent(self, prompt: str, config: dict):
        if not self.last_messages:
            self.log_console.append_log("⚠ No conversation history to continue", "warn")
            return
        
        self.call_count += 1
        self.log_console.append_log(f"▶ CONTINUE #{self.call_count}: {prompt}", "start")
        
        continue_worker = ContinueWorker(prompt, self.engine, config, self.last_messages)
        continue_worker.log_signal.connect(self.log_console.append_log)
        continue_worker.grid_signal.connect(self._on_grid_update)
        continue_worker.done_signal.connect(self._on_agent_done)
        continue_worker.error_signal.connect(self._on_agent_error)
        continue_worker.start()

        self.btn_execute.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.progress.setVisible(True)
        self.agent_label.setText("◉ AGENT RUNNING")
        self.agent_label.setStyleSheet(f"color: {PALETTE['yellow']}; font-size: 10px; font-family: monospace;")
        self.cmd_input.clear()

    def _stop_agent(self):
        if self.agent_worker and self.agent_worker.isRunning():
            self.agent_worker.terminate()
            self.agent_worker.wait(1000)
            self.log_console.append_log("■ AGENT TERMINATED BY USER", "warn")
            self._agent_reset()

    def _on_grid_update(self):
        self._refresh_grid()
        self._update_sheet_combo()

    def _on_agent_done(self, msg: str):
        if self.memory and self.agent_worker:
            for m in self.agent_worker._messages:
                if m.get("role") in ["user", "assistant"]:
                    self.memory.add_message(m["role"], m.get("content", "")[:500])
        
        if "PAUSED" in msg or "continue" in msg.lower():
            self.log_console.append_log(f"◈ {msg}", "warn")
            if self.agent_worker:
                self.last_messages = self.agent_worker._messages
                self.is_paused = True
            self.agent_label.setText("◉ AGENT PAUSED - SAY 'CONTINUE'")
            self.agent_label.setStyleSheet(f"color: {PALETTE['warn']}; font-size: 10px; font-family: monospace;")
            self.btn_execute.setEnabled(True)
            self.btn_stop.setEnabled(False)
            self.progress.setVisible(False)
            self.log_console.append_log("💡 Type 'continue' in command box to resume", "info")
            return
        
        self.log_console.append_log(f"✓ TASK COMPLETE: {msg}", "success")
        self.log_console.append_log("─" * 80, "divider")
        
        if self.memory:
            import asyncio
            try:
                loop = asyncio.get_event_loop()
                if loop.is_running():
                    asyncio.ensure_future(self.memory.end_session(None))
                else:
                    loop.run_until_complete(self.memory.end_session(None))
            except:
                pass
        
        self._agent_reset()
        self._refresh_grid()
        self._update_sheet_combo()
        self.node_selector.update_stats(calls=self.call_count)

    def _on_agent_error(self, err: str):
        self.log_console.append_log(f"✗ AGENT ERROR: {err}", "error")
        self._agent_reset()

    def _agent_reset(self):
        self.btn_execute.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.progress.setVisible(False)
        self.agent_label.setText("◉ AGENT READY")
        self.agent_label.setStyleSheet(f"color: {PALETTE['green']}; font-size: 10px; font-family: monospace;")
        self.is_paused = False


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("SynthGrid")
    app.setApplicationVersion("1.0.0")
    
    # Dark fusion palette
    app.setStyle("Fusion")
    palette = QPalette()
    dark = QColor(PALETTE["bg0"])
    palette.setColor(QPalette.ColorRole.Window, dark)
    palette.setColor(QPalette.ColorRole.WindowText, QColor(PALETTE["text"]))
    palette.setColor(QPalette.ColorRole.Base, QColor(PALETTE["bg1"]))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor(PALETTE["bg2"]))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor(PALETTE["bg2"]))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor(PALETTE["text"]))
    palette.setColor(QPalette.ColorRole.Text, QColor(PALETTE["text"]))
    palette.setColor(QPalette.ColorRole.Button, QColor(PALETTE["bg2"]))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(PALETTE["accent"]))
    palette.setColor(QPalette.ColorRole.Highlight, QColor(PALETTE["accent"]))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(PALETTE["bg0"]))
    app.setPalette(palette)

    window = MainWindow()
    window.show()
    
    print(f"""
╔══════════════════════════════════════════════════════════╗
║  SYNTHGRID LAUNCHED                                      ║
║  ◈ GUI: Active                                           ║
║  ◈ Excel Engine: openpyxl ready                         ║
║  ◈ API keys saved to ~/.config/synthgrid/keys.json       ║
║  ◈ Set GROQ/OPENAI/ANTHROPIC_API_KEY env var or use GUI ║
║  ◈ Requires Ollama running for local inference          ║
╚══════════════════════════════════════════════════════════╝""")
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
